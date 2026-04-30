"""
Migração one-shot do Excel existente para o schema Rosatom-aware.

Estratégia: Opção 2 (Híbrida)
─────────────────────────────────────────────────────────────
1. DETERMINÍSTICO (regra) → ~880 bids viram 🔴 Baixa sem chamar API:
   - Áreas legacy "Geração", "Transmissão", "Distribuição", "Serviços Gerais"
     SEM keyword nuclear → 🔴 Baixa, frente "—"
   - "TI/Sistemas" SEM keyword nuclear → 🔴 Baixa, frente "—"
   - Justificativa: "Migração heurística — fora de escopo Rosatom (área legacy: X)"

2. API (Claude Sonnet) → ~110 bids ambíguos:
   - Áreas legacy "Nuclear", "Obras/Civil", "Material/Insumo", "Equipamento", "Outro"
     OU qualquer linha com keyword nuclear → enviado em lotes para o classificador novo
   - Reaproveita construir_prompt() de cfe_monitor.py

3. Tags estratégicas (CENTENA, Caldas, ...) detectadas para TODOS os 990 bids.

Output: o mesmo CFE_Monitor_Consolidado.xlsx é regravado com Relevância/Frente/Tags
preenchidos. Cols Revisão/Observação/Erro Class./Relev. Correta são preservadas.
Custo estimado: ~$0,12 (4 lotes de 30 bids no Claude Sonnet 4).

Uso:
    python migrar_excel_rosatom.py [--dry-run] [--excel CFE_Monitor_Consolidado.xlsx]

--dry-run: imprime a distribuição esperada (heurística + ambíguos) sem chamar API
           nem gravar o Excel.
"""
from __future__ import annotations
import argparse, os, sys, time
from pathlib import Path
from collections import Counter

# Reutiliza definições do monitor — TAG_KEYWORDS, KEYWORDS_*_LEGACY, COLUNAS, COL_*,
# detectar_tags(), construir_prompt(), AREAS_LISTA, FRENTES_VALIDAS.
import cfe_monitor as cm
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

ARQUIVO_PADRAO = "CFE_Monitor_Consolidado.xlsx"

# Áreas legacy que NÃO precisam da API — vão direto para 🔴 Baixa (a menos
# que tenham keyword nuclear)
AREAS_LEGACY_DETERMINISTICAS = {
    "Geração", "Transmissão", "Distribuição",
    "Serviços Gerais", "TI/Sistemas",
}

# Áreas legacy que SEMPRE precisam da API (caso ambíguo)
AREAS_LEGACY_AMBIGUAS = {
    "Nuclear", "Obras/Civil", "Outro",
}


def tem_keyword_nuclear(desc: str) -> bool:
    desc_low = (desc or "").lower()
    if any(kw in desc_low for kw in cm.KEYWORDS_NAO_NUCLEAR_LEGACY):
        return False
    return any(kw in desc_low for kw in cm.KEYWORDS_NUCLEAR_LEGACY)


def classificar_heuristica(area_legacy: str, descricao: str) -> tuple[str, str, str] | None:
    """
    Retorna (relevancia, frente, justificativa) se conseguir decidir sem API.
    None se for ambíguo e precisar de API.
    """
    # Se tem keyword nuclear, sempre ambíguo (vai pra API)
    if tem_keyword_nuclear(descricao):
        return None
    # Áreas claramente comerciais/genéricas → Rosatom não disputa
    if area_legacy in AREAS_LEGACY_DETERMINISTICAS:
        return (
            "🔴 Baixa",
            "—",
            f"Migração heurística — fora de escopo Rosatom (legacy: {area_legacy})",
        )
    # Outras áreas legacy ou já reclassificadas (🟢/🟡/🔴) ficam ambíguas
    return None


def ja_tem_schema_novo(relev_atual: str) -> bool:
    """True se a coluna 6 já estiver com relevância nova (🟢/🟡/🔴)."""
    return relev_atual.startswith(("🟢", "🟡", "🔴"))


CACHE_API = "migracao_api_cache.json"


def chamar_api_lote(procs_ambiguos: list[dict]) -> dict:
    """Chama a API em lotes de 30. Retorna dict {numero: {relevancia, frente, justificativa}}.

    Cacheia em disco (migracao_api_cache.json) — se rodar de novo, pula bids já
    classificados. Permite recuperar de erros pós-API sem pagar de novo.
    """
    import anthropic, httpx, json, re
    LOTE = 30

    # Carrega cache de execução anterior, se existir
    resultado: dict = {}
    if Path(CACHE_API).exists():
        try:
            resultado = json.loads(Path(CACHE_API).read_text(encoding="utf-8"))
            print(f"  Cache: {len(resultado)} bids já classificados em run anterior — vou pular esses.")
        except Exception:
            resultado = {}

    pendentes = [p for p in procs_ambiguos if p.get("numero") not in resultado]
    if not pendentes:
        print("  Todos os bids já estão no cache — nenhuma chamada API necessária.")
        return resultado
    print(f"  Pendentes para API: {len(pendentes)} bids.")

    client = anthropic.Anthropic(http_client=httpx.Client(verify=False))

    for inicio in range(0, len(pendentes), LOTE):
        lote = pendentes[inicio : inicio + LOTE]
        print(f"  API: lote {inicio//LOTE+1} ({len(lote)} bids)...")
        prompt = cm.construir_prompt(lote)
        for tentativa in range(1, 4):
            try:
                msg = client.messages.create(
                    model="claude-sonnet-4-20250514",
                    max_tokens=8192,
                    messages=[{"role": "user", "content": prompt}],
                )
                break
            except Exception as e:
                if "overloaded" in str(e).lower() and tentativa < 3:
                    time.sleep(30 * tentativa)
                else:
                    raise

        txt = re.sub(r"^```(?:json)?|```$", "", msg.content[0].text.strip(), flags=re.MULTILINE).strip()
        try:
            arr = json.loads(txt)
            for c in arr:
                idx = c.get("indice")
                if idx is None or idx >= len(lote):
                    continue
                num = lote[idx].get("numero", "")
                relev = c.get("relevancia", "🟡 Média")
                frente = c.get("frente", "—")
                if relev not in cm.AREAS_LISTA:
                    relev = "🟡 Média"
                if frente not in cm.FRENTES_VALIDAS:
                    frente = "—"
                if relev == "🔴 Baixa":
                    frente = "—"
                resultado[num] = {
                    "relevancia": relev,
                    "frente": frente,
                    "justificativa": c.get("justificativa", ""),
                }
            # Persiste cache após cada lote
            Path(CACHE_API).write_text(json.dumps(resultado, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as e:
            print(f"  ⚠ erro JSON no lote {inicio//LOTE+1}: {e}")
            print(f"  resposta: {txt[:200]!r}")
    return resultado


def estilizar_relev(c, relev):
    cor = cm.AREAS_CORES.get(relev, "#94A3B8").lstrip("#")
    c.fill = PatternFill("solid", fgColor=cor)
    c.font = Font(bold=True, size=9, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")


def estilizar_frente(c, frente):
    cor = cm.TIPO_CORES_XL.get(frente, "757575")
    c.fill = PatternFill("solid", fgColor=cor)
    c.font = Font(bold=True, size=9, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default=ARQUIVO_PADRAO)
    ap.add_argument("--dry-run", action="store_true",
                    help="Não chama API nem grava — só mostra distribuição")
    args = ap.parse_args()

    if not Path(args.excel).exists():
        print(f"ERRO: {args.excel} não encontrado.")
        sys.exit(1)

    print("=" * 60)
    print("  MIGRAÇÃO ROSATOM-AWARE")
    print(f"  Excel: {args.excel}")
    print(f"  Dry-run: {args.dry_run}")
    print("=" * 60)

    wb = load_workbook(args.excel)
    if "Base Geral" not in wb.sheetnames:
        print("ERRO: planilha 'Base Geral' não encontrada.")
        sys.exit(1)
    ws = wb["Base Geral"]

    # Coleta linhas: idx_excel -> dados relevantes
    linhas = []
    for r in range(4, ws.max_row + 1):
        num   = ws.cell(row=r, column=cm.COL_NUM).value
        if not num:
            continue
        relev = str(ws.cell(row=r, column=cm.COL_AREA).value or "")
        frent = str(ws.cell(row=r, column=cm.COL_TIPO).value or "")
        desc  = str(ws.cell(row=r, column=5).value or "")
        tipo_proc    = str(ws.cell(row=r, column=8).value or "")
        tipo_contrat = str(ws.cell(row=r, column=9).value or "")
        origem       = str(ws.cell(row=r, column=cm.COL_ORIGEM).value or "")
        linhas.append({
            "row_excel": r, "numero": str(num).strip(),
            "area_legacy": relev, "tipo_legacy": frent,
            "descripcion": desc, "tipo_proc": tipo_proc,
            "tipo_contrat": tipo_contrat, "origem": origem,
        })

    print(f"\nTotal de bids no Excel: {len(linhas)}")

    # Decide o que vai pra heurística vs API
    determinados = []   # já decidido sem API
    ambiguos     = []   # precisa API
    ja_migrados  = 0    # já está no schema novo

    for ln in linhas:
        if ja_tem_schema_novo(ln["area_legacy"]):
            ja_migrados += 1
            continue
        h = classificar_heuristica(ln["area_legacy"], ln["descripcion"])
        if h:
            determinados.append((ln, *h))
        else:
            ambiguos.append(ln)

    print(f"  Já no schema novo : {ja_migrados}")
    print(f"  Heurística (skip) : {len(determinados)} → 🔴 Baixa direto")
    print(f"  Ambíguos (API)    : {len(ambiguos)}")
    custo_estimado = (len(ambiguos) / 30 + 0.5) * 0.03  # ~$0,03 por lote
    print(f"  Custo estimado    : ~${custo_estimado:.2f}")

    if args.dry_run:
        # Mostra distribuição das áreas legacy entre os ambíguos
        cnt = Counter(ln["area_legacy"] for ln in ambiguos)
        print("\n  Distribuição das áreas LEGACY entre ambíguos:")
        for a, n in cnt.most_common():
            print(f"    {a:<22}: {n}")
        print("\n[dry-run] Nenhuma chamada feita, Excel não modificado.")
        return

    if not os.getenv("ANTHROPIC_API_KEY"):
        print("ERRO: ANTHROPIC_API_KEY não configurada.")
        sys.exit(1)

    # Chama API para os ambíguos
    api_results: dict = {}
    if ambiguos:
        # Adapta para o formato esperado por construir_prompt
        procs_p_api = [
            {
                "numero": ln["numero"],
                "origem": ln["origem"],
                "descripcion": ln["descripcion"],
                "tipo_proc": ln["tipo_proc"],
                "tipo_contrat": ln["tipo_contrat"],
            }
            for ln in ambiguos
        ]
        api_results = chamar_api_lote(procs_p_api)
        print(f"  API: {len(api_results)} respostas recebidas.")

    # Aplica resultados ao Excel
    aplicados_h = 0
    aplicados_a = 0
    cnt_relev   = Counter()
    cnt_frente  = Counter()
    cnt_tags    = Counter()

    for ln, *_ignored in []:
        pass  # placeholder

    # 1) Heurísticos
    for ln, relev, frente, justif in determinados:
        r = ln["row_excel"]
        ws.cell(row=r, column=cm.COL_AREA).value = relev
        ws.cell(row=r, column=cm.COL_TIPO).value = frente
        ws.cell(row=r, column=15).value = justif  # col 15 = Justificativa
        estilizar_relev(ws.cell(row=r, column=cm.COL_AREA), relev)
        estilizar_frente(ws.cell(row=r, column=cm.COL_TIPO), frente)
        cnt_relev[relev]  += 1
        cnt_frente[frente] += 1
        aplicados_h += 1

    # 2) Ambíguos via API
    for ln in ambiguos:
        r = ln["row_excel"]
        res = api_results.get(ln["numero"])
        if not res:
            # API falhou para esta linha — marca 🟡 Média para revisão manual
            relev, frente, justif = "🟡 Média", "—", "Migração — API não retornou; revisar manualmente"
        else:
            relev  = res["relevancia"]
            frente = res["frente"]
            justif = res["justificativa"]
        ws.cell(row=r, column=cm.COL_AREA).value = relev
        ws.cell(row=r, column=cm.COL_TIPO).value = frente
        ws.cell(row=r, column=15).value = justif
        estilizar_relev(ws.cell(row=r, column=cm.COL_AREA), relev)
        estilizar_frente(ws.cell(row=r, column=cm.COL_TIPO), frente)
        cnt_relev[relev]   += 1
        cnt_frente[frente] += 1
        aplicados_a += 1

    # 3) Tags para TODAS as linhas (offline)
    for ln in linhas:
        r = ln["row_excel"]
        desc_full = ln["descripcion"] + " " + ln["tipo_contrat"]
        tags = cm.detectar_tags(desc_full)
        if tags:
            tags_str = ", ".join(tags)
            ws.cell(row=r, column=cm.COL_TAGS).value = tags_str
            for t in tags:
                cnt_tags[t] += 1

    # 4) Renomeia headers (col 6, 7, 21, adiciona 22 se faltando)
    novos_headers = {
        cm.COL_AREA:          ("Relevância",    14),
        cm.COL_TIPO:          ("Frente",        18),
        cm.COL_AREA_CORRETA:  ("Relev. Correta",18),
        cm.COL_TAGS:          ("Tags",          22),
    }
    for col, (titulo, larg) in novos_headers.items():
        c = ws.cell(row=3, column=col, value=titulo)
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="3949AB")
        c.alignment = Alignment(horizontal="center", vertical="center")
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = larg

    wb.save(args.excel)

    print()
    print("=" * 60)
    print("  RESUMO")
    print("=" * 60)
    print(f"  Heurística aplicada : {aplicados_h}")
    print(f"  API     aplicada    : {aplicados_a}")
    print(f"  Já migrados (skip)  : {ja_migrados}")
    print()
    print("  Distribuição final por Relevância:")
    for r, n in cnt_relev.most_common():
        print(f"    {r:<10}: {n}")
    print()
    print("  Top 5 frentes:")
    for f, n in cnt_frente.most_common(5):
        print(f"    {f:<14}: {n}")
    print()
    if cnt_tags:
        print("  Tags detectadas (top 10):")
        for t, n in cnt_tags.most_common(10):
            print(f"    {t:<16}: {n}")
    print()
    print(f"✓ Excel salvo: {args.excel}")
    print("Próximo passo: rode `python validar_golden_set.py` para sanity-check.")


if __name__ == "__main__":
    main()
