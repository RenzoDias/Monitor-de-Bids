"""
CFE Monitor v5
==============
- Excel: apenas Base Geral (dados puros, sem gráficos)
- HTML:  dashboard interativo completo com gráficos, filtros e tabela
- Nuclear sempre em VERMELHO
- Upsert inteligente por número de processo

Requisitos:
    pip install requests beautifulsoup4 openpyxl anthropic httpx

Uso:
    python cfe_monitor.py
    python cfe_monitor.py --ini 2026-04-01 --fim 2026-04-08
"""

import os, sys, logging, requests, json, re, argparse, httpx
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ─────────────────────────────────────────────────────────────
#  LOGGING — saída pra stdout (capturado pelo GitHub Actions)
# ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    stream=sys.stdout,
)
from bs4 import BeautifulSoup
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import anthropic

# ─────────────────────────────────────────────────────────────
#  CONFIGURAÇÕES
# ─────────────────────────────────────────────────────────────
BASE_URL            = "https://msc.cfe.mx/Aplicaciones/NCFE/Concursos"
INDEX_URL           = f"{BASE_URL}/"
BUSCA_URL           = f"{BASE_URL}/Procedure/getProcBusqueda"
ARQUIVO_EXCEL       = "CFE_Monitor_Consolidado.xlsx"
ARQUIVO_HTML        = "CFE_Dashboard.html"
ORIGEM_PADRAO       = "CFE"

# ── Eletronuclear — API PNCP (pública, sem autenticação) ──────
ELETRONUCLEAR_UASG   = "910847"
COMPRASNET_URL       = "https://comprasnet.gov.br/ConsultaLicitacoes/ConsLicitacao_Relacao.asp"
ELETRONUCLEAR_ORIGEM = "Eletronuclear"

# ── INB — sistemas.inb.gov.br ────────────────────────────────
INB_URL_FORM   = "https://sistemas.inb.gov.br/leiacessoconsulta/consulta/ConsultarLicitacao.aspx"
INB_ORIGEM     = "INB"

# ── CDTN — Comprasnet via PNCP ───────────────────────────────
CDTN_UASG   = "113205"           # UASG do CDTN/CNEN
CDTN_ORIGEM = "CDTN"

# ── NASA — Nucleoeléctrica Argentina S.A. ────────────────────
NASA_URL    = "https://www.na-sa.com.ar/proveedores/home/licitaciones/vigentes"
NASA_ORIGEM = "NASA"

# ── Dioxitek ─────────────────────────────────────────────────
DIOXITEK_URL    = "https://dioxitek.com.ar/transparencia-activa/compras-y-contrataciones"
DIOXITEK_ORIGEM = "Dioxitek"

# ── Mercado Público Chile — CCHEN ────────────────────────────
CCHEN_URL      = "https://www.mercadopublico.cl/BuscarLicitacion/Home/Buscar"
CCHEN_CODIGO   = "7219"   # Comisión Chilena de Energía Nuclear
CCHEN_ORIGEM   = "CCHEN"

# ── IAEA — UNGM ──────────────────────────────────────────────
IAEA_URL    = "https://www.ungm.org/Public/Notice/Search"
IAEA_CODIGO = 44       # IAEA agency code no UNGM
IAEA_ORIGEM = "IAEA"

# ── Lookback por fonte (em dias) ──────────────────────────────
# Fontes esparsas precisam de janela mais ampla pra capturar bids.
# Sem override (None) = usa ini/fim padrão da CLI/calcular_intervalo.
# Bids já na base são deduplicados via procs_existentes (numero).
LOOKBACK_DIAS_FONTE = {
    "CFE":            None,   # alta volume — janela padrão
    "Eletronuclear":  None,   # média volume — janela padrão
    "INB":            None,   # média volume — janela padrão
    "NASA":           None,   # alta volume — janela padrão
    "Dioxitek":       None,   # média volume — janela padrão
    "CDTN":           60,     # esparso — 60 dias
    "CCHEN":          30,     # esparso — 30 dias
    "IAEA":           30,     # esparso — 30 dias
}

# ── GitHub Pages ───────────────────────────────────────────────
GITHUB_TOKEN  = os.getenv("GITHUB_TOKEN", "")
GITHUB_USER   = "RenzoDias"
GITHUB_REPO   = "Monitor-de-Bids"
GITHUB_BRANCH = "main"

# ─────────────────────────────────────────────────────────────
#  SCHEMA ROSATOM-AWARE (substitui Area/Tipo do classificador antigo)
#  - col "Area"  do Excel agora contém RELEVÂNCIA  (🟢 Alta / 🟡 Média / 🔴 Baixa)
#  - col "Tipo"  do Excel agora contém FRENTE      (TVEL / ASE / Uranium One / ...)
#  - col "Tags"  (nova, col 22) contém tags estratégicas detectadas por keywords
# ─────────────────────────────────────────────────────────────
AREAS_CORES = {
    "🟢 Alta"  : "#22C55E",  # verde — disputar
    "🟡 Média" : "#F59E0B",  # âmbar — revisar manualmente
    "🔴 Baixa" : "#94A3B8",  # cinza-azulado — Rosatom não disputa
}
AREAS_LISTA = list(AREAS_CORES.keys())

# Cores das frentes Rosatom (equivalente ao TIPO_CORES_XL antigo)
TIPO_CORES_XL = {
    "TVEL"         : "1565C0",   # combustível, lítio-7, zircônio
    "ASE"          : "6A1B9A",   # construção de usinas
    "Uranium One"  : "E65100",   # mineração de urânio
    "Metal Tech"   : "1B5E20",   # titânio, materiais estratégicos
    "RWM"          : "B71C1C",   # rejeitos radioativos / CENTENA
    "Healthcare"   : "00838F",   # medicina nuclear
    "NovaWind"     : "558B2F",   # eólica
    "Múltiplas"    : "37474F",
    "—"            : "475569",
    "Outro"        : "757575",   # fallback
}

# Tags estratégicas — detectadas por keywords na descrição (PT/ES/EN)
TAG_KEYWORDS = {
    "CENTENA"        : ["centena", "depósito final de rejeitos", "deposito final de rejeitos",
                        "depósito de rejeitos radioativos", "deposito de rejeitos radioativos", "rwdf"],
    "Caldas"         : ["caldas", "poços de caldas", "pocos de caldas"],
    "Caetité"        : ["caetité", "caetite"],
    "Santa Quitéria" : ["santa quitéria", "santa quiteria", "itataia"],
    "Angra 3"        : ["angra 3", "angra iii", "angra-3"],
    "SMR"            : ["smr ", " smr", "small modular reactor", "reator modular", "reactor modular"],
    "Urânio"         : ["urânio", "uranio", "u3o8", "yellowcake", "ucp ", "concentrado de uranio",
                        "concentrado de urânio"],
    "Titânio"        : ["titânio", "titanio", "titanium"],
    "Lítio-7"        : ["lítio-7", "litio-7", "lithium-7", "lítio enriquecido", "litio enriquecido"],
    "Zircônio"       : ["zircônio", "zirconio", "zirconium", "zircaloy"],
}

# Frentes válidas que o Claude pode emitir (validação de output)
FRENTES_VALIDAS = {"TVEL", "ASE", "Uranium One", "Metal Tech", "RWM",
                   "Healthcare", "NovaWind", "Múltiplas", "—"}

# ─── Keywords antigas (LEGADO — usadas pela migração para classificar
# determinísticamente os ~880 bids antigos como 🔴 Baixa quando claramente
# fora de escopo Rosatom). NÃO usadas pelo classificador novo. ──────────
KEYWORDS_NUCLEAR_LEGACY = [
    "nuclear", "combustible nuclear", "uranio", "radioactiv", "radiactiv",
    "radioativ", "radiativ", "fonte radioativ", "fontes radioativ", "fonte de radia",
    "laguna verde", "cnlv", "fision", "fisión", "combustível nuclear", "urânio",
    "planta nuclear", "central nuclear", "reactor nuclear", "reator nuclear",
]
KEYWORDS_NAO_NUCLEAR_LEGACY = [
    "amortiguamiento", "subestacion", "subestación", "banco de reactores",
    "reactor shunt", "reactor en derivacion", "compensacion reactiva", "compensación reactiva",
]

COLUNAS = [
    ("Status",          14),  # 1
    ("Data Pub.",        12),  # 2
    ("Origem",           10),  # 3
    ("Número",           26),  # 4
    ("Descrição",        55),  # 5
    ("Relevância",       14),  # 6  ← 🟢 Alta / 🟡 Média / 🔴 Baixa (era "Área")
    ("Frente",           18),  # 7  ← TVEL/ASE/Uranium One/... (era "Tipo")
    ("Tipo Proc.",       18),  # 8
    ("Contratação",      18),  # 9
    ("Estado",           14),  # 10
    ("Prazo Submissão",  22),  # 11
    ("Julgamento",       22),  # 12
    ("Valor (MXN)",      16),  # 13
    ("Entidade",         22),  # 14
    ("Justificativa",    40),  # 15
    ("Atualizado em",    16),  # 16
    ("Campos Alterados", 35),  # 17
    ("Revisão",          16),  # 18  ← vem de revisoes.csv
    ("Observação",       50),  # 19  ← vem de revisoes.csv
    ("Erro Class.",      14),  # 20  ← vem de revisoes.csv (🔴 Sim / 🟢 Não)
    ("Relev. Correta",   18),  # 21  ← vem de revisoes.csv (preenchido se erro)
    ("Tags",             22),  # 22  ← detectadas auto por keywords (CENTENA, Caldas, ...)
]
COL_STATUS=1;COL_DATA=2;COL_ORIGEM=3;COL_NUM=4
# COL_AREA → agora "Relevância", COL_TIPO → agora "Frente". Mantidos os nomes
# de variável para minimizar diff e preservar referências em montar_linha/estilo_linha.
COL_AREA=6;COL_TIPO=7;COL_PRAZO=11
COL_ATUALIZADO=16;COL_ALTERADOS=17
COL_REVISAO=18;COL_OBSERVACAO=19
COL_ERRO_CLAS=20;COL_AREA_CORRETA=21
COL_TAGS=22

# Colunas ignoradas na comparação (mudam sempre, não indicam mudança real)
COLUNAS_IGNORAR_COMPARACAO = {
    COL_STATUS,    # status muda intencionalmente
    13,            # Valor (MXN) — formato inconsistente
    15,            # Justificativa — gerada pelo Claude
    16,            # Atualizado em — timestamp sempre novo
    17,            # Campos Alterados — calculado aqui
    COL_REVISAO,        # Revisão — vem de revisoes.csv
    COL_OBSERVACAO,     # Observação — vem de revisoes.csv
    COL_ERRO_CLAS,      # Erro Classificação — vem de revisoes.csv
    COL_AREA_CORRETA,   # Relev. Correta — vem de revisoes.csv
    COL_TAGS,           # Tags — detectadas auto a cada run; não indica mudança da fonte
}

# Opções de revisão (para referência)
OPCOES_REVISAO = ["✔ Seguido", "✘ Não seguido", "👁 Em análise", "⏸ Aguardando"]
OPCOES_ERRO_CLAS = ["🔴 Sim - área errada", "🟢 Não - classificação ok"]
ARQUIVO_REVISOES = "revisoes.csv"

# Nomes das colunas para exibir no campo "Campos Alterados"
NOMES_COLUNAS = {i+1: nome for i,(nome,_) in enumerate(COLUNAS)}

COR_HEADER="1A237E";COR_SUBHEAD="283593";COR_ALT="E8EAF6"

def borda():
    s=Side(style="thin",color="CCCCCC")
    return Border(left=s,right=s,top=s,bottom=s)
def hdr(c,cor=COR_HEADER,sz=13):
    c.font=Font(bold=True,size=sz,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor=cor)
    c.alignment=Alignment(horizontal="center",vertical="center")
    c.border=borda()
def chdr(c,cor="3949AB"):
    c.font=Font(bold=True,size=9,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor=cor)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border=borda()

# ─────────────────────────────────────────────────────────────
#  1. SESSÃO E TOKEN
# ─────────────────────────────────────────────────────────────
def criar_sessao():
    s=requests.Session()
    s.headers.update({"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/146.0.0.0 Safari/537.36","Accept-Language":"es-MX,es;q=0.9"})
    return s

def obter_token(session):
    print("Obtendo token...")
    r=session.get(INDEX_URL,timeout=30); r.raise_for_status()
    soup=BeautifulSoup(r.text,"html.parser")
    inp=soup.find("input",{"name":"__RequestVerificationToken"})
    if inp: return inp["value"]
    meta=soup.find("meta",{"name":"RequestVerificationToken"})
    if meta: return meta.get("content","")
    raise RuntimeError("Token não encontrado.")

# ─────────────────────────────────────────────────────────────
#  2. INTERVALO
# ─────────────────────────────────────────────────────────────
def calcular_intervalo():
    hoje=date.today(); dias=3 if hoje.weekday()==0 else 1
    ini=hoje-timedelta(days=dias)
    print(f"{'Segunda' if hoje.weekday()==0 else 'Dia normal'} → {ini} a {hoje}")
    return ini.isoformat(),hoje.isoformat()

# ─────────────────────────────────────────────────────────────
#  3. BUSCA
# ─────────────────────────────────────────────────────────────
def buscar(session,token,ini,fim):
    print(f"Buscando {ini} a {fim}...")
    r=session.post(BUSCA_URL,data={
        "__RequestVerificationToken":token,
        "TipoProcedimientoClave":"","TipoContratacionClave":"",
        "IdEntidadFederativa":"0","Numero":"","Descripcion":"",
        "EstadoProcedimientoContratacionClave":"0",
        "FechaPublicacion":"","FechaPublicacionIni":ini,"FechaPublicacionFin":fim,
        "TestigoSocial":"2","idCaracterProcedimiento":"0","Modalidad":"0",
    },headers={"Content-Type":"application/x-www-form-urlencoded; charset=UTF-8",
               "X-Requested-With":"XMLHttpRequest","Referer":INDEX_URL,"Origin":"https://msc.cfe.mx"},
    timeout=60); r.raise_for_status()
    try: dados=r.json()
    except: dados=[]
    return normalizar(dados)

def normalizar(dados):
    if isinstance(dados,dict):
        for k in ["data","Data","procedimientos"]:
            if k in dados: dados=dados[k]; break
    procs=[]
    for item in (dados if isinstance(dados,list) else []):
        if isinstance(item,dict):
            procs.append({
                "numero":item.get("Numero",item.get("numero","")),
                "descripcion":item.get("Descripcion",item.get("descripcion","")),
                "tipo_proc":item.get("TipoProcedimiento",item.get("tipoProcedimiento","")),
                "tipo_contrat":item.get("TipoContratacion",item.get("tipoContratacion","")),
                "estado":item.get("EstadoProcedimiento",item.get("estadoProcedimiento","")),
                "fecha_pub":item.get("FechaPublicacion",item.get("fechaPublicacion","")),
                "entidad":item.get("EntidadFederativa",item.get("entidadFederativa","")),
                "prazo_sub":"",   # preenchido pelo detalhe
                "julgamento":"",  # preenchido pelo detalhe
                "monto":item.get("MONTO",item.get("Monto","")),
                "id_interno":str(item.get("Id",item.get("id",""))),
                "origem":ORIGEM_PADRAO,"area":"","tipo":"","justificativa":"",
            })
    return procs

def normalizar_data(val):
    s=str(val or "").strip()
    if not s: return ""
    m=re.search(r'/Date\((\d+)',s)
    if m:
        dt=datetime.fromtimestamp(int(m.group(1))/1000,tz=timezone.utc)
        return dt.strftime("%Y-%m-%d")
    m=re.match(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',s)
    if m: return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    if re.match(r'\d{4}-\d{2}-\d{2}',s): return s[:10]
    return s[:10]


# ─────────────────────────────────────────────────────────────
#  3b. BUSCA DE DATAS NO DETALHE
# ─────────────────────────────────────────────────────────────
DETALHE_URL = f"{BASE_URL}/Procedure/Details"

def buscar_datas_detalhe(session, id_interno: str) -> dict:
    """Abre a página de detalhe via POST com o Id numérico e extrai datas."""
    resultado = {"prazo_sub": "", "julgamento": ""}
    if not id_interno: return resultado

    tentativas = [
        ("POST", DETALHE_URL, {"id": id_interno}),
        ("POST", DETALHE_URL, {"Id": id_interno}),
        ("GET",  f"{DETALHE_URL}?id={id_interno}", None),
    ]

    for metodo, url, payload in tentativas:
        try:
            if metodo == "POST":
                resp = session.post(url, data=payload,
                    headers={"Referer": INDEX_URL, "Origin": "https://msc.cfe.mx"},
                    timeout=20)
            else:
                resp = session.get(url, timeout=20)

            if resp.status_code != 200: continue
            soup = BeautifulSoup(resp.text, "html.parser")
            texto = soup.get_text(" ", strip=True)
            if "Submiss" not in texto and "submiss" not in texto: continue

            # Procura seção "Submissão de propostas" → Data de término
            for elem in soup.find_all(["h2","h3","h4","div","section"]):
                txt = elem.get_text(strip=True).lower()
                if "submiss" in txt and "proposta" in txt:
                    tabela = elem.find_next("table")
                    if tabela:
                        for tr in tabela.find_all("tr"):
                            cols = [td.get_text(strip=True) for td in tr.find_all("td")]
                            if len(cols) >= 2:
                                label = cols[0].lower()
                                if "término" in label or "termino" in label or "fim" in label:
                                    resultado["prazo_sub"] = normalizar_data(cols[1])
                                elif "início" in label or "inicio" in label:
                                    pass  # não precisamos do início

            # Procura julgamento / fallo
            for elem in soup.find_all(["h2","h3","h4","div","section"]):
                txt = elem.get_text(strip=True).lower()
                if "julgamento" in txt or "falha" in txt or "fallo" in txt:
                    tabela = elem.find_next("table")
                    if tabela:
                        for tr in tabela.find_all("tr"):
                            cols = [td.get_text(strip=True) for td in tr.find_all("td")]
                            if len(cols) >= 2 and re.search(r"\d{2}/\d{2}/\d{4}", cols[1]):
                                resultado["julgamento"] = normalizar_data(cols[1])
                                break

            if resultado["prazo_sub"] or resultado["julgamento"]:
                return resultado

        except Exception:
            continue

    return resultado


def buscar_todos_detalhes(session, procs: list) -> list:
    total = len(procs)
    print(f"Buscando prazos nos detalhes ({total} procedimentos)...")
    achou = 0
    for i, p in enumerate(procs, 1):
        r = buscar_datas_detalhe(session, p.get("id_interno",""))
        p["prazo_sub"]  = r["prazo_sub"]
        p["julgamento"] = r["julgamento"]
        if r["prazo_sub"]: achou += 1
        if i % 10 == 0 or i == total:
            print(f"  [{i}/{total}] prazos encontrados: {achou}")
        import time; time.sleep(0.4)
    return procs

# ─────────────────────────────────────────────────────────────
#  4. CLAUDE
# ─────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────
#  BUSCA ELETRONUCLEAR — Comprasnet
# ─────────────────────────────────────────────────────────────
def buscar_eletronuclear(session, ini: str, fim: str) -> list:
    """
    Busca licitações da Eletronuclear via Comprasnet (UASG 910847).
    Usa POST para ConsLicitacao_Relacao.asp com intervalo de datas.
    Faz scraping do HTML retornado.
    """
    from bs4 import BeautifulSoup

    print(f"Buscando Eletronuclear (Comprasnet) de {ini} a {fim}...")

    # Converte datas de YYYY-MM-DD para DD/MM/YYYY
    def fmt(d): return d[8:10] + '%2F' + d[5:7] + '%2F' + d[:4]

    payload = (
        f"numprp=&dt_publ_ini={fmt(ini)}&dt_publ_fim={fmt(fim)}"
        f"&txtObjeto=&chkModalidade=1&chkModalidade=2&chkModalidade=3"
        f"&chkModalidade=20&chkModalidade=5&chkModalidade=99"
        f"&chkTodos=-1&optTpPesqMat=M&optTpPesqServ=S"
        f"&txtlstUasg={ELETRONUCLEAR_UASG}"
        f"&txtlstUf=&txtlstMunicipio=&txtlstModalidade=&txtlstTpPregao="
        f"&txtlstConcorrencia=&txtlstGrpMaterial=&txtlstClasMaterial="
        f"&txtlstMaterial=&txtlstGrpServico=&txtlstServico=&Origem=F"
    )

    headers = {
        "Content-Type" : "application/x-www-form-urlencoded",
        "Referer"      : "https://comprasnet.gov.br/ConsultaLicitacoes/ConsLicitacao_Filtro.asp",
        "User-Agent"   : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Origin"       : "https://comprasnet.gov.br",
    }

    try:
        r = session.post(
            COMPRASNET_URL,
            data=payload,
            headers=headers,
            timeout=30,
            verify=False,
        )
        r.raise_for_status()
        # Força encoding correto — Comprasnet usa windows-1252
        r.encoding = r.apparent_encoding or "windows-1252"
        html = r.text
    except Exception as e:
        print(f"  Aviso Comprasnet: {e}")
        return []

    soup = BeautifulSoup(html, "html.parser")
    procs = []

    import re

    # Estrutura do Comprasnet: cada licitação começa com "ELETRONUCLEAR S.A"
    # em negrito, seguido dos campos em linhas com <b>Campo:</b> valor
    # Estratégia: divide o HTML em blocos por ocorrência de "ELETRONUCLEAR"

    # Pega todo o texto da página dividido por tags <b>
    # e processa bloco a bloco
    texto_completo = soup.get_text(separator="|SEP|")
    # Divide nos blocos de cada licitação
    blocos_txt = re.split(r'ELETRONUCLEAR S\.?A\.?', texto_completo, flags=re.IGNORECASE)

    for bloco_txt in blocos_txt[1:]:  # pula o primeiro (antes da primeira ocorrência)
        try:
            linhas = [l.strip() for l in bloco_txt.replace("|SEP|", "\n").split("\n") if l.strip()]

            numero = ""
            objeto = ""
            data_pub = ""
            abertura = ""
            modalidade = "Licitação"

            prox_e_data_pub = False
            prox_e_abertura = False

            for linha in linhas:
                l = linha.lower()

                # Número: "Pregăo Eletrônico Nş 90038/2026" (encoding corrompido)
                # Aceita qualquer variação de "Preg", "N" seguido de símbolo e número
                m = re.search(r'preg\w*\s+eletr\w*\s+n\S?\s*(\d+/\d{4})', linha, re.IGNORECASE)
                if m:
                    numero = m.group(1)
                    modalidade = "Pregão Eletrônico"
                else:
                    m2 = re.search(r'(concorr\w*|dispensa|inexig\w*)\s+n\S?\s*(\d+/\d{4})', linha, re.IGNORECASE)
                    if m2:
                        numero = m2.group(2)
                        tipo = m2.group(1).upper()
                        if "CONC" in tipo: modalidade = "Concorrência"
                        elif "DISP" in tipo: modalidade = "Dispensa"
                        elif "INEX" in tipo: modalidade = "Inexigibilidade"

                # Objeto — linha começa com "Objeto:"
                if l.startswith("objeto:"):
                    obj = linha[7:].strip()
                    if obj.lower().startswith("objeto:"):
                        obj = obj[7:].strip()
                    # Remove "Pregăo Eletrônico - " do início se duplicado
                    obj = re.sub(r'^preg\w*\s+eletr\w*\s*[-–]\s*', '', obj, flags=re.IGNORECASE).strip()
                    if obj: objeto = obj

                # Data publicação: label e data podem estar na mesma linha ou em linhas separadas
                if "edital a partir" in l:
                    m = re.search(r'(\d{2}/\d{2}/\d{4})', linha)
                    if m:
                        data_pub = normalizar_data(m.group(1))
                    else:
                        prox_e_data_pub = True  # data vem na próxima linha
                elif prox_e_data_pub:
                    m = re.search(r'(\d{2}/\d{2}/\d{4})', linha)
                    if m: data_pub = normalizar_data(m.group(1))
                    prox_e_data_pub = False

                # Fallback data: "Entrega da Proposta"
                if "entrega da proposta" in l and not data_pub:
                    m = re.search(r'(\d{2}/\d{2}/\d{4})', linha)
                    if m: data_pub = normalizar_data(m.group(1))

                # Abertura
                if "abertura da proposta" in l:
                    m = re.search(r'(\d{2}/\d{2}/\d{4})', linha)
                    if m:
                        abertura = normalizar_data(m.group(1))
                    else:
                        prox_e_abertura = True
                elif prox_e_abertura:
                    m = re.search(r'(\d{2}/\d{2}/\d{4})', linha)
                    if m: abertura = normalizar_data(m.group(1))
                    prox_e_abertura = False

                if "nova pesquisa" in l:
                    break

            if not objeto and not numero:
                continue

            procs.append({
                "numero"       : f"ETN-{numero}" if numero else f"ETN-{len(procs)+1}",
                "descripcion"  : objeto,
                "tipo_proc"    : modalidade,
                "tipo_contrat" : "",
                "estado"       : "Publicado",
                "fecha_pub"    : data_pub,
                "entidad"      : "Eletronuclear",
                "prazo_sub"    : "",
                "julgamento"   : abertura,
                "monto"        : "",
                "id_interno"   : numero,
                "origem"       : ELETRONUCLEAR_ORIGEM,
                "area"         : "",
                "tipo"         : "",
                "justificativa": "",
            })
        except Exception:
            continue

    if procs:
        print(f"  Total Eletronuclear: {len(procs)} licitações encontradas.")
    else:
        print(f"  Nenhuma licitação da Eletronuclear no período {ini} a {fim}.")
    return procs


# ─────────────────────────────────────────────────────────────
#  BUSCA INB — sistemas.inb.gov.br
# ─────────────────────────────────────────────────────────────
def buscar_inb(session, ini: str, fim: str) -> list:
    """
    Busca licitações da INB via portal sistemas.inb.gov.br.
    Estratégia:
      1. GET na página para obter __VIEWSTATE e __EVENTVALIDATION
      2. POST com ano atual, situação=Todas, exportar=XML
      3. Filtra no Python pelo intervalo de datas
    """
    from bs4 import BeautifulSoup
    import re

    ano = ini[:4]  # ano da data inicial
    print(f"Buscando INB de {ini} a {fim} (ano {ano})...")

    # ── 1. GET para obter tokens ASP.NET ──────────────────────
    try:
        r = session.get(INB_URL_FORM, timeout=30, verify=False)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        vs  = soup.find("input", {"id": "__VIEWSTATE"})
        vsg = soup.find("input", {"id": "__VIEWSTATEGENERATOR"})
        ev  = soup.find("input", {"id": "__EVENTVALIDATION"})

        viewstate  = vs["value"]  if vs  else ""
        vsgenerator= vsg["value"] if vsg else ""
        eventval   = ev["value"]  if ev  else ""
    except Exception as e:
        print(f"  Aviso INB (GET): {e}")
        return []

    # ── 2. POST para exportar XML ─────────────────────────────
    payload = {
        "__VIEWSTATE"        : viewstate,
        "__VIEWSTATEGENERATOR": vsgenerator,
        "__EVENTVALIDATION"  : eventval,
        "ctl00$ContentPlaceHolder1$Componente$cboAno"          : ano,
        "ctl00$ContentPlaceHolder1$Componente$cboModalidade"   : "0",
        "ctl00$ContentPlaceHolder1$Componente$cboSituacao"     : "1",  # Em andamento
        "ctl00$ContentPlaceHolder1$Componente$txtNumeroLicitacao": "",
        "ctl00$ContentPlaceHolder1$Componente$txtObjeto"       : "",
        "ctl00$ContentPlaceHolder1$Componente$cboFormato"      : "XML",
        "ctl00$ContentPlaceHolder1$Componente$btnExportar"     : "Exportar",
    }
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Referer"     : INB_URL_FORM,
        "User-Agent"  : "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Origin"      : "https://sistemas.inb.gov.br",
    }
    try:
        r2 = session.post(INB_URL_FORM, data=payload, headers=headers,
                          timeout=60, verify=False)
        r2.raise_for_status()
        xml_text = r2.text

    except Exception as e:
        print(f"  Aviso INB (POST): {e}")
        import traceback; traceback.print_exc()
        return []

    # ── 3. Parse do XML ───────────────────────────────────────
    # Tenta XML; se falhar, tenta HTML (o site às vezes retorna HTML mesmo pedindo XML)
    procs = []
    ini_dt = ini  # YYYY-MM-DD
    fim_dt = fim

    try:
        import xml.etree.ElementTree as ET
        root = ET.fromstring(xml_text.encode("utf-8"))

        for item in root.iter():
            tag = item.tag.lower().replace("{", "").split("}")[-1]
            if tag != "licitacao":
                continue
            _extrair_inb(item, procs, ini_dt, fim_dt)

    except ET.ParseError:
        soup2 = BeautifulSoup(xml_text, "html.parser")
        _parse_inb_html(soup2, procs, ini_dt, fim_dt)

    if procs:
        print(f"  Total INB: {len(procs)} licitações encontradas.")
    else:
        print(f"  Nenhuma licitação da INB no período {ini} a {fim}.")
    return procs


def _extrair_inb(item, procs, ini_dt, fim_dt):
    """Extrai campos de um elemento XML da INB.
    Tags confirmadas: NumeroProcessoLicitacao, NumeroProcessoAquisicao,
    Objeto, Modalidade, DataHoraPublicacao, DataHoraAbertura, DataHoraDisputa,
    LocalAbertura, NumeroBB, Descricao, Anexos/Anexo/Caminho
    """
    # Busca apenas filhos DIRETOS da Licitacao (evita pegar tags dentro de Anexos)
    def g(tag):
        for child in item:
            if child.tag.lower() == tag.lower():
                return (child.text or "").strip()
        return ""

    numero     = g("NumeroProcessoLicitacao")
    objeto     = g("Objeto").strip()
    modalidade = g("Modalidade") or "Licitação"
    situacao   = g("Situacao") or "Em andamento"
    local_ab   = g("LocalAbertura")
    unidade    = g("UnidadeAdministrativa")

    # Datas diretas da Licitacao (não dentro de Anexo)
    data_pub_raw = g("DataHoraPublicacao")   # ex: "10/04/2026"
    abertura_raw = g("DataHoraAbertura")     # ex: "27/04/2026 10:00"

    data_pub = normalizar_data(data_pub_raw[:10]) if data_pub_raw else ""
    abertura = normalizar_data(abertura_raw[:10]) if abertura_raw else ""

    if not numero and not objeto:
        return

    # Normaliza modalidade
    mod_up = modalidade.upper()
    if "PREG" in mod_up and "ELETR" in mod_up:
        modalidade = "Pregão Eletrônico"
    elif "CONCORR" in mod_up and "FECHA" in mod_up:
        modalidade = "Concorrência Fechada"
    elif "CONCORR" in mod_up:
        modalidade = "Concorrência"
    elif "DISPENSA" in mod_up:
        modalidade = "Dispensa"
    elif "INEXIG" in mod_up:
        modalidade = "Inexigibilidade"

    procs.append({
        "numero"       : f"INB-{numero}",
        "descripcion"  : objeto,
        "tipo_proc"    : modalidade,
        "tipo_contrat" : "",
        "estado"       : situacao,
        "fecha_pub"    : data_pub,
        "entidad"      : unidade or "INB",
        "prazo_sub"    : "",
        "julgamento"   : abertura,
        "monto"        : "",
        "id_interno"   : numero,
        "origem"       : INB_ORIGEM,
        "area"         : "",
        "tipo"         : "",
        "justificativa": "",
    })


def _parse_inb_html(soup, procs, ini_dt, fim_dt):
    """
    Extrai licitações do HTML/XML retornado pelo INB.
    Estrutura real observada:
      91.028/2026  PROC  OBJETO  EM ANDAMENTO  ...  DATA_PUB  COMPRASNET  ...  ABERTURA  DATA_PUB
    Padrão: número/ano seguido de processo, objeto, situação, documentos com datas, e
    ao final: ABERTURA_DATA  MODALIDADE  ABERTURA_DATA  DATA_PUB  SISTEMA
    """
    import re

    texto = soup.get_text(separator=" | ")

    # Divide em blocos por número de licitação (padrão: 91.XXX/AAAA ou X.XXX/AAAA)
    blocos = re.split(r'(?=(?:PE\s+GCONT\.F\s+)?\d+\.\d+/20\d{2}(?:\s|\|))', texto)

    for bloco in blocos:
        try:
            bloco = bloco.strip()
            if not bloco or len(bloco) < 30:
                continue

            # Extrai número da licitação
            m_num = re.search(r'((?:PE\s+GCONT\.F\s+)?[\d]+\.[\d]+/20\d{2})', bloco)
            if not m_num:
                continue
            numero_raw = m_num.group(1).strip()
            # Normaliza: remove "PE GCONT.F " se presente
            numero = re.sub(r'PE\s+GCONT\.F\s+', '', numero_raw).strip()

            # Objeto: texto longo em maiúsculas após o número do processo
            # Procura padrão PROCESSO seguido de texto em maiúsculas
            m_obj = re.search(
                r'[A-Z]{2,}\.[A-Z]+-\d{4}/\d{2}/\d{4}[^A-Z]*([A-ZÁÀÂÃÉÈÊÍÏÓÔÕÚÇÑ][^|]{30,}?)(?:\s*\|\s*EM ANDAMENTO|\s*\|\s*ENCERRADO)',
                bloco, re.IGNORECASE
            )
            objeto = m_obj.group(1).strip() if m_obj else ""

            # Se não achou, tenta pegar texto longo entre o número do processo e "EM ANDAMENTO"
            if not objeto:
                m_obj2 = re.search(r'\d{4}/\d{2}/\d{4}\s*\|?\s*([^|]{40,}?)\s*\|?\s*EM ANDAMENTO', bloco, re.IGNORECASE)
                if m_obj2:
                    objeto = m_obj2.group(1).strip()

            # Modalidade
            if re.search(r'PREG[ÃA]O ELETR[ÔO]NICO', bloco, re.IGNORECASE):
                modalidade = "Pregão Eletrônico"
            elif re.search(r'CONCORR[ÊE]NCIA FECHADA', bloco, re.IGNORECASE):
                modalidade = "Concorrência Fechada"
            elif re.search(r'CONCORR[ÊE]NCIA', bloco, re.IGNORECASE):
                modalidade = "Concorrência"
            elif re.search(r'DISPENSA', bloco, re.IGNORECASE):
                modalidade = "Dispensa"
            elif re.search(r'INEXIGIBILIDADE', bloco, re.IGNORECASE):
                modalidade = "Inexigibilidade"
            else:
                modalidade = "Licitação"

            # Data de abertura: vem antes de "10:00"
            abertura = ""
            m_ab = re.search(r'(\d{2}/\d{2}/20\d{2})\s+\d{2}:\d{2}', bloco)
            if m_ab:
                abertura = normalizar_data(m_ab.group(1))

            # Data de publicação: label explícita "Publicado em: DD/MM/AAAA"
            data_pub = ""
            m_pub = re.search(r'Publicado\s+em\s*[:\|]?\s*(\d{2}/\d{2}/20\d{2})', bloco, re.IGNORECASE)
            if m_pub:
                data_pub = normalizar_data(m_pub.group(1))

            # Fallback: última data do bloco que não seja abertura
            if not data_pub:
                todas_datas = re.findall(r'\b(\d{2}/\d{2}/20\d{2})\b', bloco)
                for d in reversed(todas_datas):
                    nd = normalizar_data(d)
                    if nd != abertura:
                        data_pub = nd
                        break
                if not data_pub and todas_datas:
                    data_pub = normalizar_data(todas_datas[0])

            if not objeto and not numero:
                continue

            # Filtra pelo intervalo de datas
            if data_pub and not (ini_dt <= data_pub <= fim_dt):
                continue

            procs.append({
                "numero"       : f"INB-{numero}",
                "descripcion"  : objeto,
                "tipo_proc"    : modalidade,
                "tipo_contrat" : "",
                "estado"       : "Em andamento",
                "fecha_pub"    : data_pub,
                "entidad"      : "INB",
                "prazo_sub"    : "",
                "julgamento"   : abertura,
                "monto"        : "",
                "id_interno"   : numero,
                "origem"       : INB_ORIGEM,
                "area"         : "",
                "tipo"         : "",
                "justificativa": "",
            })
        except Exception:
            continue



# ─────────────────────────────────────────────────────────────
#  BUSCA CDTN — Comprasnet (UASG 113205)
# ─────────────────────────────────────────────────────────────
def buscar_cdtn(session, ini: str, fim: str) -> list:
    """
    Busca licitações do CDTN via Comprasnet (UASG 113205).
    Reutiliza a mesma lógica da Eletronuclear.
    """
    from bs4 import BeautifulSoup
    import re

    print(f"Buscando CDTN (Comprasnet) de {ini} a {fim}...")

    def fmt(d): return d[8:10] + '%2F' + d[5:7] + '%2F' + d[:4]

    payload = (
        f"numprp=&dt_publ_ini={fmt(ini)}&dt_publ_fim={fmt(fim)}"
        f"&txtObjeto=&chkModalidade=1&chkModalidade=2&chkModalidade=3"
        f"&chkModalidade=20&chkModalidade=5&chkModalidade=99"
        f"&chkTodos=-1&optTpPesqMat=M&optTpPesqServ=S"
        f"&txtlstUasg={CDTN_UASG}"
        f"&txtlstUf=&txtlstMunicipio=&txtlstModalidade=&txtlstTpPregao="
        f"&txtlstConcorrencia=&txtlstGrpMaterial=&txtlstClasMaterial="
        f"&txtlstMaterial=&txtlstGrpServico=&txtlstServico=&Origem=F"
    )
    headers = {
        "Content-Type" : "application/x-www-form-urlencoded",
        "Referer"      : "https://comprasnet.gov.br/ConsultaLicitacoes/ConsLicitacao_Filtro.asp",
        "User-Agent"   : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Origin"       : "https://comprasnet.gov.br",
    }
    try:
        r = session.post(
            COMPRASNET_URL,
            data=payload,
            headers=headers,
            timeout=30,
            verify=False,
        )
        r.raise_for_status()
        r.encoding = r.apparent_encoding or "windows-1252"
        html = r.text
    except Exception as e:
        print(f"  Aviso CDTN Comprasnet: {e}")
        return []

    soup = BeautifulSoup(html, "html.parser")
    procs = []
    texto_completo = soup.get_text(separator="|SEP|")

    # O Comprasnet retorna o nome do órgão no topo — busca pelo nome CDTN
    blocos_txt = re.split(r'CENTRO\s+DE\s+DESENV\S*\s+DA\s+TECNOLOGIA|CDTN|SAE-CNEN', texto_completo, flags=re.IGNORECASE)

    for bloco_txt in blocos_txt[1:]:
        try:
            linhas = [l.strip() for l in bloco_txt.replace("|SEP|", "\n").split("\n") if l.strip()]

            numero     = ""
            objeto     = ""
            data_pub   = ""
            abertura   = ""
            modalidade = "Licitação"
            prox_e_data_pub = False
            prox_e_abertura = False

            for linha in linhas:
                l = linha.lower()

                m = re.search(r'preg\w*\s+eletr\w*\s+n\S?\s*(\d+/\d{4})', linha, re.IGNORECASE)
                if m:
                    numero = m.group(1); modalidade = "Pregão Eletrônico"
                else:
                    m2 = re.search(r'(concorr\w*|dispensa|inexig\w*)\s+n\S?\s*(\d+/\d{4})', linha, re.IGNORECASE)
                    if m2:
                        numero = m2.group(2)
                        tipo = m2.group(1).upper()
                        if "CONC" in tipo: modalidade = "Concorrência"
                        elif "DISP" in tipo: modalidade = "Dispensa"
                        elif "INEX" in tipo: modalidade = "Inexigibilidade"

                if l.startswith("objeto:"):
                    obj = linha[7:].strip()
                    if obj.lower().startswith("objeto:"): obj = obj[7:].strip()
                    obj = re.sub(r'^preg\w*\s+eletr\w*\s*[-–]\s*', '', obj, flags=re.IGNORECASE).strip()
                    if obj: objeto = obj

                if "edital a partir" in l:
                    m = re.search(r'(\d{2}/\d{2}/\d{4})', linha)
                    if m: data_pub = normalizar_data(m.group(1))
                    else: prox_e_data_pub = True
                elif prox_e_data_pub:
                    m = re.search(r'(\d{2}/\d{2}/\d{4})', linha)
                    if m: data_pub = normalizar_data(m.group(1))
                    prox_e_data_pub = False

                if "entrega da proposta" in l and not data_pub:
                    m = re.search(r'(\d{2}/\d{2}/\d{4})', linha)
                    if m: data_pub = normalizar_data(m.group(1))

                if "abertura da proposta" in l:
                    m = re.search(r'(\d{2}/\d{2}/\d{4})', linha)
                    if m: abertura = normalizar_data(m.group(1))
                    else: prox_e_abertura = True
                elif prox_e_abertura:
                    m = re.search(r'(\d{2}/\d{2}/\d{4})', linha)
                    if m: abertura = normalizar_data(m.group(1))
                    prox_e_abertura = False

                if "nova pesquisa" in l: break

            if not objeto and not numero: continue

            procs.append({
                "numero"       : f"CDTN-{numero}" if numero else f"CDTN-{len(procs)+1}",
                "descripcion"  : objeto,
                "tipo_proc"    : modalidade,
                "tipo_contrat" : "",
                "estado"       : "Publicado",
                "fecha_pub"    : data_pub,
                "entidad"      : "CDTN/CNEN",
                "prazo_sub"    : "",
                "julgamento"   : abertura,
                "monto"        : "",
                "id_interno"   : numero,
                "origem"       : CDTN_ORIGEM,
                "area"         : "",
                "tipo"         : "",
                "justificativa": "",
            })
        except Exception:
            continue

    if procs:
        print(f"  Total CDTN: {len(procs)} licitações encontradas.")
    else:
        print(f"  Nenhuma licitação do CDTN no período {ini} a {fim}.")
    return procs



# ─────────────────────────────────────────────────────────────
#  BUSCA NASA — Nucleoeléctrica Argentina S.A.
# ─────────────────────────────────────────────────────────────
def buscar_nasa(session, ini: str, fim: str) -> list:
    """
    Busca licitações da NASA via GET simples.
    Toda a página carrega de uma vez — sem filtro por data no servidor.
    As 3 abas (Sede Central, CNA I-II, CNE) estão no mesmo HTML.
    Filtra no Python pelo intervalo de datas (Fecha de Apertura).
    """
    from bs4 import BeautifulSoup
    import re

    print(f"Buscando NASA (Nucleoeléctrica Argentina) de {ini} a {fim}...")

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept"    : "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer"   : "https://www.na-sa.com.ar/proveedores/home",
    }
    try:
        r = session.get(NASA_URL, headers=headers, timeout=30, verify=False)
        r.raise_for_status()
        r.encoding = "utf-8"
        html = r.text
    except Exception as e:
        print(f"  Aviso NASA: {e}")
        return []

    soup = BeautifulSoup(html, "html.parser")
    procs = []

    # Mapa de abas para nomes de unidade
    ABAS = {
        "sede": "Sede Central",
        "cna" : "CNA I-II",
        "cne" : "CNE",
    }

    for aba_id, unidade in ABAS.items():
        div = soup.find("div", {"id": aba_id})
        if not div:
            continue
        for tr in div.find_all("tr"):
            tds = tr.find_all("td")
            if len(tds) < 4:
                continue

            expediente  = tds[0].get_text(strip=True)
            descricao   = tds[1].get_text(strip=True)
            tipo        = tds[2].get_text(strip=True)
            fecha_texto = tds[3].get_text(strip=True)  # "2026-04-17 a las 11:00:00 horas"

            if not expediente or not descricao:
                continue

            # Extrai data de abertura: "2026-04-17 a las 11:00:00 horas"
            m = re.search(r'(\d{4}-\d{2}-\d{2})', fecha_texto)
            abertura = m.group(1) if m else ""  # já em YYYY-MM-DD

            # NASA não tem data de publicação — usa abertura como proxy
            # Filtra: só traz licitações com abertura dentro do intervalo
            # OBS: como é "vigentes", traz tudo que ainda está aberto
            # Deixamos sem filtro de data para não perder nada
            # O upsert cuida de novos vs existentes

            # NASA não expõe data de publicação — usa data de hoje
            # Para licitações já na base, o upsert preserva a data original
            from datetime import date as _date
            data_pub = _date.today().strftime("%Y-%m-%d")

            numero_id = f"{aba_id.upper()}-{expediente}"

            # Descarta licitações com abertura já vencida
            if abertura and abertura < _date.today().strftime("%Y-%m-%d"):
                continue

            procs.append({
                "numero"       : f"NASA-{expediente}",
                "descripcion"  : descricao,
                "tipo_proc"    : tipo,
                "tipo_contrat" : "",
                "estado"       : "Vigente",
                "fecha_pub"    : data_pub,
                "entidad"      : f"NA-SA / {unidade}",
                "prazo_sub"    : "",
                "julgamento"   : abertura,
                "monto"        : "",
                "id_interno"   : expediente,
                "origem"       : NASA_ORIGEM,
                "area"         : "",
                "tipo"         : "",
                "justificativa": "",
            })

    if procs:
        print(f"  Total NASA: {len(procs)} licitações encontradas.")
    else:
        print(f"  Nenhuma licitação da NASA encontrada.")
    return procs



# ─────────────────────────────────────────────────────────────
#  BUSCA DIOXITEK
# ─────────────────────────────────────────────────────────────
def buscar_dioxitek(session, ini: str, fim: str) -> list:
    """
    Busca licitações da Dioxitek (Argentina) via GET simples.
    A página carrega tudo de uma vez em HTML estático.
    Estrutura: blocos com h6 como título e parágrafos com detalhes.
    """
    from bs4 import BeautifulSoup
    import re
    from datetime import date as _date

    print(f"Buscando Dioxitek de {ini} a {fim}...")

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Accept"    : "text/html,application/xhtml+xml",
    }
    try:
        r = session.get(DIOXITEK_URL, headers=headers, timeout=30, verify=False)
        r.raise_for_status()
        r.encoding = "utf-8"
        html = r.text
    except Exception as e:
        print(f"  Aviso Dioxitek: {e}")
        return []

    soup = BeautifulSoup(html, "html.parser")
    procs = []
    vistos = set()  # evita duplicatas de circulares do mesmo expediente

    # Cada licitacao aparece como h6 seguido de h6 (objeto) e paragrafos
    h6s = soup.find_all("h6")

    i = 0
    while i < len(h6s):
        h6 = h6s[i]
        titulo = h6.get_text(strip=True)

        # Procura número de expediente: EX-AAAA-XXXXXXXX
        m_exp = re.search(r'(EX-\d{4}-\d+(?:-+\S+)*)', titulo, re.IGNORECASE)
        if not m_exp:
            i += 1
            continue

        expediente = re.search(r'EX-\d{4}-\d+', m_exp.group(1)).group(0)

        # Evita duplicatas — pega só a primeira ocorrência do expediente
        if expediente in vistos:
            i += 1
            continue
        vistos.add(expediente)

        # Tipo de licitação (início do título)
        tipo_match = re.match(r'^(Concurso de Precios|Licitaci[oó]n P[uú]blica|Licitaci[oó]n Privada|Contrataci[oó]n Directa|Concurso)', titulo, re.IGNORECASE)
        tipo = tipo_match.group(1) if tipo_match else "Concurso de Precios"

        # Objeto: h6 seguinte geralmente tem "OBJETO:"
        objeto = ""
        if i + 1 < len(h6s):
            prox = h6s[i+1].get_text(strip=True)
            if "OBJETO" in prox.upper() or (not re.search(r'EX-\d{4}', prox) and len(prox) > 10):
                objeto = re.sub(r'(?i)^objeto\s*:\s*', '', prox).strip()

        # Se não achou no próximo h6, busca nos parágrafos seguintes
        if not objeto:
            for sib in h6.find_next_siblings():
                if sib.name == "h6":
                    txt = sib.get_text(strip=True)
                    if "OBJETO" in txt.upper():
                        objeto = re.sub(r'(?i)^objeto\s*:\s*', '', txt).strip()
                    break
                if sib.name in ("p", "ul", "li"):
                    txt = sib.get_text(strip=True)
                    if "OBJETO" in txt.upper():
                        objeto = re.sub(r'(?i)^objeto\s*:\s*', '', txt).strip()
                        break

        # Data de abertura: busca nos parágrafos seguintes
        abertura = ""
        texto_bloco = ""
        for sib in h6.find_next_siblings():
            if sib.name == "h6" and re.search(r'EX-\d{4}', sib.get_text()):
                break  # próxima licitação
            texto_bloco += " " + sib.get_text(separator=" ", strip=True)

        # Padrões de data: "24 de abril del 2026", "06 de abril de 2026"
        MESES = {"enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
                 "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12}
        m_data = re.search(
            r'(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+del?\s+(\d{4})',
            texto_bloco, re.IGNORECASE
        )
        if m_data:
            dia  = int(m_data.group(1))
            mes  = MESES.get(m_data.group(2).lower(), 1)
            ano  = int(m_data.group(3))
            abertura = f"{ano:04d}-{mes:02d}-{dia:02d}"

        data_pub = _date.today().strftime("%Y-%m-%d")

        if not objeto:
            objeto = titulo

        # Descarta licitações com abertura já vencida
        if abertura and abertura < _date.today().strftime("%Y-%m-%d"):
            i += 1
            continue

        procs.append({
            "numero"       : f"DIOXITEK-{expediente}",
            "descripcion"  : objeto,
            "tipo_proc"    : tipo,
            "tipo_contrat" : "",
            "estado"       : "Vigente",
            "fecha_pub"    : data_pub,
            "entidad"      : "Dioxitek S.A.",
            "prazo_sub"    : abertura,
            "julgamento"   : abertura,
            "monto"        : "",
            "id_interno"   : expediente,
            "origem"       : DIOXITEK_ORIGEM,
            "area"         : "",
            "tipo"         : "",
            "justificativa": "",
        })
        i += 1

    if procs:
        print(f"  Total Dioxitek: {len(procs)} licitações encontradas.")
    else:
        print(f"  Nenhuma licitação da Dioxitek encontrada.")
    return procs



# ─────────────────────────────────────────────────────────────
#  BUSCA CCHEN — Mercado Público Chile
# ─────────────────────────────────────────────────────────────
def buscar_cchen(session, ini: str, fim: str) -> list:
    """
    Busca licitações da CCHEN via Mercado Público Chile.
    POST retorna HTML com divs .lic-bloq-wrap.
    Busca sem filtro de datas para pegar todas publicadas/vigentes
    e filtra no Python pelo intervalo ini-fim.
    """
    import json as _json
    from bs4 import BeautifulSoup
    import re

    print(f"Buscando CCHEN (Mercado Público Chile) de {ini} a {fim}...")

    headers = {
        "Content-Type"    : "application/json",
        "User-Agent"      : "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Referer"         : "https://www.mercadopublico.cl/Home/BusquedaLicitacion",
        "Origin"          : "https://www.mercadopublico.cl",
        "Accept"          : "*/*",
        "X-Requested-With": "XMLHttpRequest",
    }

    procs  = []
    pagina = 0

    while True:
        payload = {
            "textoBusqueda"        : "",
            "idEstado"             : "5",
            "codigoRegion"         : "-1",
            "idTipoLicitacion"     : "-1",
            "fechaInicio"          : None,
            "fechaFin"             : None,
            "registrosPorPagina"   : "50",
            "idTipoFecha"          : [],
            "idOrden"              : "1",
            "compradores"          : [CCHEN_CODIGO],
            "garantias"            : None,
            "rubros"               : [],
            "proveedores"          : [],
            "montoEstimadoTipo"    : [0],
            "esPublicoMontoEstimado": None,
            "pagina"               : pagina,
        }

        try:
            r = session.post(
                CCHEN_URL,
                data=_json.dumps(payload),
                headers=headers,
                timeout=30,
                verify=False,
            )
            r.raise_for_status()
            html = r.text
        except Exception as e:
            print(f"  Aviso CCHEN pág {pagina}: {e}")
            break

        soup = BeautifulSoup(html, "html.parser")
        blocos = soup.find_all("div", class_="lic-bloq-wrap")

        if not blocos:
            break

        for bloco in blocos:
            # ID Licitación
            id_el = bloco.find("div", class_="id-licitacion")
            numero = id_el.find("span").get_text(strip=True) if id_el else ""

            # Título / objeto
            h2 = bloco.find("h2")
            titulo = h2.get_text(strip=True) if h2 else ""

            # Descrição (parágrafo após h2)
            p_desc = bloco.find("div", class_="lic-block-body")
            ps = p_desc.find_all("p", class_="text-weight-light") if p_desc else []
            descricao = ps[0].get_text(strip=True) if ps else titulo

            # Estado
            estado_el = bloco.find("span", class_="estado-texto")
            estado = estado_el.get_text(strip=True) if estado_el else "Publicada"

            # Tipo (sigla LE, LP, etc.)
            tipo_el = bloco.find("div", class_="estado-lic")
            tipo_sigla = tipo_el.find("span").get_text(strip=True) if tipo_el else ""
            TIPOS_SIGLA = {"LE":"Licitación Pública LE", "LP":"Licitación Pública LP",
                           "LR":"Licitación Privada", "LS":"Licitación de Servicios"}
            tipo = TIPOS_SIGLA.get(tipo_sigla, tipo_sigla or "Licitación")

            # Datas e monto — busca dentro de margin-bottom-md row
            data_pub = ""
            cierre   = ""
            monto    = ""
            for div in bloco.find_all("div", class_="margin-bottom-md"):
                texto = div.get_text(separator="|")
                # Fecha publicación
                m = re.search(r'Fecha de publicaci[oó]n\|(\d{2}/\d{2}/\d{4})', texto)
                if m: data_pub = normalizar_data(m.group(1))
                # Fecha cierre
                m = re.search(r'Fecha de cierre\|(\d{2}/\d{2}/\d{4})', texto)
                if m: cierre = normalizar_data(m.group(1))
                # Monto
                spans = div.find_all("span", class_="campo-numerico-punto-coma")
                if spans: monto = spans[0].get_text(strip=True)

            # Filtra pelo intervalo de datas de publicação
            if data_pub and not (ini <= data_pub <= fim):
                continue

            if not numero and not titulo:
                continue

            procs.append({
                "numero"       : f"CCHEN-{numero}",
                "descripcion"  : descricao or titulo,
                "tipo_proc"    : tipo,
                "tipo_contrat" : "",
                "estado"       : estado,
                "fecha_pub"    : data_pub,
                "entidad"      : "CCHEN",
                "prazo_sub"    : cierre,
                "julgamento"   : cierre,
                "monto"        : monto,
                "id_interno"   : numero,
                "origem"       : CCHEN_ORIGEM,
                "area"         : "",
                "tipo"         : "",
                "justificativa": "",
            })

        # Verifica total para paginação
        total_el = soup.find("input", {"id": "hdnTotalPresupuestoPublico"})
        total = int(total_el["value"]) if total_el else 0
        if len(procs) >= total or len(blocos) < 50:
            break
        pagina += 1

    if procs:
        print(f"  Total CCHEN: {len(procs)} licitações encontradas.")
    else:
        print(f"  Nenhuma licitação da CCHEN no período {ini} a {fim}.")
    return procs



# ─────────────────────────────────────────────────────────────
#  BUSCA IAEA — UNGM (United Nations Global Marketplace)
# ─────────────────────────────────────────────────────────────
def buscar_iaea(session, ini: str, fim: str) -> list:
    """
    Busca licitações da IAEA via UNGM.
    POST retorna HTML — cada licitação é um div[data-noticeid].
    Estrutura confirmada:
      span.ungm-title           → título
      div.resultInfo1.deadline  → deadline "27-Apr-2026 17:00"
      4ª div.tableCell          → data publicação "09-Apr-2026"
      div.resultInfo1[data-description=Reference] → referência
      última div.tableCell      → país beneficiário
    Filtra por América Latina no Python.
    """
    import json as _json
    from bs4 import BeautifulSoup
    import re
    from datetime import date as _date

    print(f"Buscando IAEA (UNGM) de {ini} a {fim}...")

    MESES_EN = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
                7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
    def fmt(d):
        if not d: return ""
        y, m, dia = d[:10].split("-")
        return f"{int(dia):02d}-{MESES_EN[int(m)]}-{y}"

    def nd(s):
        """Converte DD-Mon-YYYY para YYYY-MM-DD."""
        if not s: return ""
        MES = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
               "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
        m = re.search(r"(\d{1,2})-([A-Za-z]{3})-(\d{4})", str(s))
        if m: return f"{m.group(3)}-{MES.get(m.group(2),1):02d}-{int(m.group(1)):02d}"
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", str(s))
        if m: return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        return ""

    PAISES_AL = {
        "argentina","bolivia","brazil","chile","colombia","costa rica","cuba",
        "ecuador","el salvador","guatemala","haiti","honduras","mexico",
        "nicaragua","panama","paraguay","peru","dominican republic","uruguay","venezuela",
    }

    headers = {
        "Content-Type"    : "application/json",
        "User-Agent"      : "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Referer"         : "https://www.ungm.org/Public/Notice",
        "Origin"          : "https://www.ungm.org",
        "Accept"          : "*/*",
        "X-Requested-With": "XMLHttpRequest",
    }

    hoje   = _date.today().strftime("%Y-%m-%d")
    procs  = []
    vistos = set()
    pagina = 0
    POR_PAG = 50

    while True:
        payload = {
            "PageIndex"              : pagina,
            "PageSize"               : POR_PAG,
            "Title"                  : "",
            "Description"            : "",
            "Reference"              : "",
            "PublishedFrom"          : fmt(ini),
            "PublishedTo"            : fmt(fim),
            "DeadlineFrom"           : fmt(hoje),
            "DeadlineTo"             : "",
            "Countries"              : [],
            "Agencies"               : [IAEA_CODIGO],
            "UNSPSCs"                : [],
            "NoticeTypes"            : [],
            "SortField"              : "Deadline",
            "SortAscending"          : True,
            "isPicker"               : False,
            "IsSustainable"          : False,
            "IsActive"               : True,
            "NoticeDisplayType"      : None,
            "NoticeSearchTotalLabelId": "noticeSearchTotal",
            "TypeOfCompetitions"     : [],
        }

        try:
            r = session.post(
                IAEA_URL,
                data=_json.dumps(payload),
                headers=headers,
                timeout=30,
                verify=False,
            )
            r.raise_for_status()
            html = r.text
        except Exception as e:
            print(f"  Aviso IAEA pág {pagina}: {e}")
            break

        soup  = BeautifulSoup(html, "html.parser")
        linhas = soup.find_all("div", attrs={"data-noticeid": True})

        if not linhas:
            break

        for row in linhas:
            notice_id = row.get("data-noticeid","")
            if notice_id in vistos:
                continue
            vistos.add(notice_id)

            # Título
            titulo_el = row.find("span", class_="ungm-title")
            titulo = titulo_el.get_text(strip=True) if titulo_el else ""

            # Deadline: div.resultInfo1.deadline
            dead_el = row.find("div", class_=lambda c: c and "deadline" in c)
            deadline_raw = dead_el.get_text(strip=True) if dead_el else ""
            deadline = nd(deadline_raw)

            # Todas as células
            cells = row.find_all("div", role="cell")

            # Data publicação: 4ª célula (índice 3)
            data_pub = ""
            if len(cells) > 3:
                data_pub = nd(cells[3].get_text(strip=True))

            # Referência: div com data-description="Reference"
            ref_el = row.find("div", attrs={"data-description": "Reference"})
            if ref_el:
                referencia = ref_el.get_text(strip=True)
            else:
                referencia = notice_id

            # Tipo: célula com label
            tipo = ""
            for cell in cells:
                lbl = cell.find("label")
                if lbl:
                    tipo = lbl.get_text(strip=True)
                    break

            # País beneficiário: última célula de texto simples
            pais = ""
            for cell in reversed(cells):
                txt = cell.get_text(strip=True)
                if txt and len(txt) < 60 and not any(x in txt.lower() for x in ["iaea","request","quotation","proposal","bid","interest"]):
                    pais = txt.lower()
                    break

            # Filtra AL
            if pais and not any(p in pais for p in PAISES_AL):
                continue

            if not titulo and not referencia:
                continue

            procs.append({
                "numero"       : f"IAEA-{referencia}",
                "descripcion"  : titulo,
                "tipo_proc"    : tipo,
                "tipo_contrat" : "",
                "estado"       : "Active",
                "fecha_pub"    : data_pub,
                "entidad"      : "IAEA",
                "prazo_sub"    : deadline,
                "julgamento"   : deadline,
                "monto"        : "",
                "id_interno"   : referencia,
                "origem"       : IAEA_ORIGEM,
                "area"         : "",
                "tipo"         : "",
                "justificativa": "",
            })

        # Paginação
        total_el = soup.find(id="noticeSearchTotal")
        total = 0
        if total_el:
            try: total = int(total_el.get_text(strip=True).replace(",",""))
            except: pass
        if len(vistos) >= total or len(linhas) < POR_PAG:
            break
        pagina += 1

    if procs:
        print(f"  Total IAEA: {len(procs)} licitações encontradas.")
    else:
        print(f"  Nenhuma licitação da IAEA no período {ini} a {fim}.")
    return procs


def detectar_tags(descricao: str) -> list:
    """Detecta tags estratégicas (CENTENA, Caldas, Caetité, ...) por keyword
    matching na descrição. Roda offline, sem API. Retorna lista ordenada."""
    if not descricao:
        return []
    desc = descricao.lower()
    encontradas = []
    for tag, kws in TAG_KEYWORDS.items():
        for kw in kws:
            if kw in desc:
                encontradas.append(tag)
                break
    return encontradas


def construir_prompt(procs):
    """
    Prompt do classificador Rosatom-aware.
    Persona: Analista Sênior de Compras e Desenvolvimento de Negócios da Rosatom AL.
    Schema de output: relevancia | frente | justificativa (tags são detectadas
    em paralelo no Python por detectar_tags(), não pedidas ao Claude).
    """
    linhas = [
        f"{i}|{p.get('origem','?')}|{p.get('numero','')}|{p.get('descripcion','')}|{p.get('tipo_proc','')}|{p.get('tipo_contrat','')}"
        for i, p in enumerate(procs)
    ]

    origens_presentes = set(p.get("origem","") for p in procs)
    ctx_parts = []
    if ORIGEM_PADRAO in origens_presentes:       ctx_parts.append("CFE (Comision Federal de Electricidad, México) — espanhol")
    if ELETRONUCLEAR_ORIGEM in origens_presentes: ctx_parts.append("Eletronuclear (Angra 1/2, Brasil) — português")
    if INB_ORIGEM in origens_presentes:          ctx_parts.append("INB - Indústrias Nucleares do Brasil — português")
    if CDTN_ORIGEM in origens_presentes:         ctx_parts.append("CDTN/CNEN (Brasil, pesquisa nuclear) — português")
    if NASA_ORIGEM in origens_presentes:         ctx_parts.append("NA-SA Nucleoeléctrica Argentina (Atucha/Embalse) — espanhol")
    if DIOXITEK_ORIGEM in origens_presentes:     ctx_parts.append("Dioxitek S.A. (Argentina, UO2 e Co-60) — espanhol")
    if CCHEN_ORIGEM in origens_presentes:        ctx_parts.append("CCHEN (Chile, regulação/pesquisa nuclear) — espanhol")
    if IAEA_ORIGEM in origens_presentes:         ctx_parts.append("IAEA (Viena) — inglês")
    ctx = "Origens das licitações:\n  " + "\n  ".join(ctx_parts) if ctx_parts else "Licitações do setor nuclear/energia da América Latina."

    return f"""Você é um Analista Sênior de Compras e Desenvolvimento de Negócios da Rosatom
América Latina, subsidiária da estatal russa Rosatom — a maior corporação nuclear
integrada do mundo. Sua função é avaliar licitações públicas e identificar quais
representam OPORTUNIDADE COMERCIAL para a Rosatom AL.

A Rosatom atua através de 7 frentes:
  1. TVEL — combustível nuclear, lítio-7, zircônio, isótopos estáveis
  2. ASE — construção de usinas (VVER-1000/1200, SMR), Angra 3, SMR Petrobras
  3. RWM — gestão de resíduos radioativos, descomissionamento, PROJETO CENTENA
  4. Metal Tech — titânio (esponja/lingotes), zircônio, materiais estratégicos
  5. Healthcare — isótopos médicos e industriais, medicina nuclear
  6. Uranium One/Tenex — mineração de urânio, conversão (Caldas, Caetité, Santa Quitéria)
  7. NovaWind — energia eólica (menos ativo no Brasil)

PROJETOS PRIORITÁRIOS (qualquer menção é candidato a 🟢):
  CENTENA, Caldas/Poços de Caldas, Caetité, Santa Quitéria, Angra 3, SMR.

═══ PRINCÍPIOS DE AVALIAÇÃO ═══

P1. RELEVÂNCIA, não "é nuclear?". Sua decisão é se a Rosatom AL deveria
    DISPUTAR essa licitação. Há itens nucleares-genéricos que Rosatom não vende
    e itens não-rotulados como nucleares que são oportunidade clara.

P2. INSTALAÇÃO NUCLEAR ≠ ITEM NUCLEAR. Estar em Angra, FCN, URA Caetité ou
    qualquer instalação nuclear NÃO torna o item relevante. Rosatom NÃO disputa:
    limpeza, jardinagem, vigilância, alimentação, veículos, uniformes, mobiliário,
    copos descartáveis, água mineral, postes de concreto, software ERP/CAM
    genérico, manutenção elétrica convencional, geração/transmissão/distribuição
    clássica.

P3. QUALIFICAÇÃO ESPECIAL = SINAL FORTE. Itens que exigem qualificação nuclear
    formal (Class 1E, Q-grade, ASME III, RCC-M, certificação CNEN) ou em contato
    com material radioativo são candidatos a 🟢.

P4. MINERAÇÃO DE URÂNIO É CORE. Qualquer atividade ligada a mina de urânio
    (Caetité, Caldas, Santa Quitéria) é interesse direto Uranium One/Tenex:
    estudos geológicos, perfuração, caracterização de solo, movimentação de
    minério, recuperação de tanques URA, gestão de pilha de estéril radioativo
    → 🟢.

P5. AMBIGUIDADE = 🟡 PARA REVISÃO. Quando não estiver claro se é commodity ou
    item especializado, marque 🟡 com justificativa explicando o que precisa
    ser verificado no termo de referência. O Renzo revisa pessoalmente esses casos.

═══ EXEMPLOS (few-shots dos 13 INBs revisados) ═══

[🟢 Alta]
  "Recuperação tanque TQ-6305 na Unidade de Concentração de Urânio - URA"
    → relevancia:🟢 Alta, frente:Uranium One, just:"Infra direta de mina de urânio Caetité — capability core"
  "Estudos geológicos da pilha de estéril da URA"
    → relevancia:🟢 Alta, frente:Uranium One, just:"Caracterização geológica de pilha estéril radioativa — capability TVEL/RWM"
  "Movimentação de material rochoso desmontado, minério de oportunidade"
    → relevancia:🟢 Alta, frente:Uranium One, just:"Operação direta de mineração de urânio"

[🟡 Média]
  "Usinagem de 1092 hastes de aço inoxidável austenítico"
    → relevancia:🟡 Média, frente:Metal Tech, just:"Hastes em aço inox austenítico — verificar qualificação nuclear no TR"

[🔴 Baixa — falsos positivos típicos do classificador antigo]
  "Postes e cruzetas de concreto, posto CIF Caetité"
    → relevancia:🔴 Baixa, frente:—, just:"Postes de concreto — infra comum, mesmo em Caetité"
  "Coleta, transporte e destinação final de resíduos sólidos por coprocessamento"
    → relevancia:🔴 Baixa, frente:—, just:"Resíduos não radioativos (coprocessamento) — fora de escopo"
  "Software ESPIRIT CAM — atualização e suporte"
    → relevancia:🔴 Baixa, frente:—, just:"Software CAM genérico — não é capability Rosatom"
  "Equipamentos para ampliação de cobertura de rádios na FCN"
    → relevancia:🔴 Baixa, frente:—, just:"Rede de rádios comum em instalação nuclear — fora de escopo"
  "Suporte técnico e operacional ao Horto Florestal"
    → relevancia:🔴 Baixa, frente:—, just:"Manejo de horto florestal — Serviços Gerais"
  "Fornecimento parcelado de água mineral em garrafões"
    → relevancia:🔴 Baixa, frente:—, just:"Commodity — Rosatom não disputa"
  "Fornecimento de copo descartável 200ml"
    → relevancia:🔴 Baixa, frente:—, just:"Commodity descartável — fora de escopo"
  "Conservação e limpeza nas áreas da INB Caetité"
    → relevancia:🔴 Baixa, frente:—, just:"Limpeza — Rosatom não disputa, mesmo em sítio Caetité"

═══ DADOS A CLASSIFICAR ═══

{ctx}

PROCEDIMENTOS (indice|origem|numero|descricao|tipo_proc|tipo_contrato):
{chr(10).join(linhas)}

═══ FORMATO DE OUTPUT ═══

Para cada licitação retorne JSON sem markdown:

[{{"indice":0,"relevancia":"🟢 Alta"|"🟡 Média"|"🔴 Baixa","frente":"TVEL"|"ASE"|"Uranium One"|"Metal Tech"|"RWM"|"Healthcare"|"NovaWind"|"Múltiplas"|"—","justificativa":"máx 25 palavras em linguagem de comprador"}}, ...]

Regras:
- frente = "—" quando relevancia = 🔴 Baixa
- justificativa NUNCA repete a descrição; explica POR QUE é (ou não é) oportunidade
- inclua TODOS os índices da lista, na ordem"""


def analisar(procs):
    if not procs: return procs
    import time as _time

    # Processa em lotes de 30 para evitar truncamento do JSON
    LOTE = 30
    client = anthropic.Anthropic(http_client=httpx.Client(verify=False))
    cls = {}  # indice global -> classificacao

    for inicio in range(0, len(procs), LOTE):
        lote = procs[inicio:inicio+LOTE]
        # Reindexamos o lote de 0..N mas guardamos o offset para o índice global
        print(f"Analisando procedimentos {inicio+1}–{inicio+len(lote)} de {len(procs)}...")
        prompt = construir_prompt(lote)

        for tentativa in range(1, 4):
            try:
                msg = client.messages.create(
                    model="claude-sonnet-4-20250514", max_tokens=8192,
                    messages=[{"role":"user","content":prompt}])
                break
            except Exception as e:
                if "overloaded" in str(e).lower() and tentativa < 3:
                    espera = 30 * tentativa
                    print(f"  API sobrecarregada - aguardando {espera}s (tentativa {tentativa}/3)...")
                    _time.sleep(espera)
                else:
                    raise

        txt = re.sub(r"^```(?:json)?|```$","",msg.content[0].text.strip(),flags=re.MULTILINE).strip()
        try:
            lote_cls = {c["indice"]: c for c in json.loads(txt)}
            # Converte índice local para global
            for idx_local, c in lote_cls.items():
                cls[inicio + idx_local] = c
        except Exception as e:
            print(f"  Aviso: erro ao parsear JSON do Claude no lote {inicio//LOTE+1}: {e}")
            print(f"  Resposta recebida: {txt[:200]!r}")

    for i, p in enumerate(procs):
        c = cls.get(i, {})

        # Schema novo: relevancia + frente + justificativa
        relevancia = c.get("relevancia", "🟡 Média")
        frente     = c.get("frente",     "—")
        justif     = c.get("justificativa", "")

        # Normalização defensiva — emoji pode vir sem espaço
        if relevancia and not any(relevancia.startswith(e) for e in ("🟢","🟡","🔴")):
            r_low = relevancia.lower()
            if "alta"  in r_low: relevancia = "🟢 Alta"
            elif "média" in r_low or "media" in r_low: relevancia = "🟡 Média"
            elif "baixa" in r_low: relevancia = "🔴 Baixa"
            else: relevancia = "🟡 Média"
        if relevancia not in AREAS_LISTA:
            relevancia = "🟡 Média"
        if frente not in FRENTES_VALIDAS:
            frente = "—"
        # Coerção: 🔴 Baixa sempre tem frente "—"
        if relevancia == "🔴 Baixa":
            frente = "—"

        # Tags detectadas no Python (offline) a partir da descrição + contratação
        desc_full = (p.get("descripcion","") + " " + p.get("tipo_contrat",""))
        tags = detectar_tags(desc_full)

        # As colunas físicas Excel COL_AREA/COL_TIPO foram repurposed:
        #   p["area"] = relevância,  p["tipo"] = frente
        # Mantém os nomes de chave para reduzir diff no Excel/dashboard.
        p["area"]          = relevancia
        p["tipo"]          = frente
        p["tags"]          = ", ".join(tags) if tags else ""
        p["justificativa"] = justif

    return procs

# ─────────────────────────────────────────────────────────────
#  5. EXCEL — apenas Base Geral
# ─────────────────────────────────────────────────────────────
def normalizar_val(v) -> str:
    """Normaliza valor para comparação: None/null/0/0.0 → '' ou valor limpo."""
    s = str(v or "").strip()
    # Trata strings que representam None
    if s.lower() in ("none", "null", "nan"): return ""
    # Normaliza zeros numéricos: '0', '0.0', '0,0' → '0'
    try:
        f = float(s.replace(",","."))
        if f == 0: return "0"
        # Remove decimais desnecessários: '25.0' → '25'
        return str(int(f)) if f == int(f) else f"{f:.2f}"
    except ValueError:
        pass
    return s


def montar_linha(p, status="🆕 Novo", campos_alterados=""):
    # Layout: 17 colunas geradas pela ferramenta + 4 do CSV revisões + 1 Tags = 22.
    # Cols 18-21 (Revisão/Observação/Erro Class./Relev. Correta) são preservadas
    # do Excel existente em salvar_excel(); aqui ficam vazias.
    return [
        status,                                    # 1
        normalizar_data(p.get("fecha_pub","")),    # 2
        p.get("origem",    ORIGEM_PADRAO),         # 3
        p.get("numero",    ""),                    # 4
        p.get("descripcion",""),                   # 5
        p.get("area",      "🟡 Média"),            # 6  Relevância
        p.get("tipo",      "—"),                   # 7  Frente
        p.get("tipo_proc", ""),                    # 8
        p.get("tipo_contrat",""),                  # 9
        p.get("estado",    ""),                    # 10
        normalizar_data(p.get("prazo_sub","")),    # 11
        normalizar_data(p.get("julgamento","")),   # 12
        p.get("monto",     ""),                    # 13
        p.get("entidad",   ""),                    # 14
        p.get("justificativa",""),                 # 15
        datetime.now().strftime("%d/%m/%Y %H:%M"), # 16
        campos_alterados,                          # 17
        "",                                        # 18 Revisão (do CSV)
        "",                                        # 19 Observação (do CSV)
        "",                                        # 20 Erro Class. (do CSV)
        "",                                        # 21 Relev. Correta (do CSV)
        p.get("tags",      ""),                    # 22 Tags
    ]

def estilo_linha(ws,row_n,vals,cor_bg):
    relev = str(vals[COL_AREA-1]) if len(vals)>=COL_AREA else "🟡 Média"
    frente= str(vals[COL_TIPO-1]) if len(vals)>=COL_TIPO else "—"
    alta  = relev.startswith("🟢")
    media = relev.startswith("🟡")
    for col,val in enumerate(vals[:len(COLUNAS)],1):
        c=ws.cell(row=row_n,column=col,value=val)
        c.border=borda(); c.alignment=Alignment(vertical="center",wrap_text=True); c.font=Font(size=9)
        if col==COL_STATUS:
            bg,fg=("1B5E20","FFFFFF") if "Novo" in str(val) else (("E65100","FFFFFF") if "Atual" in str(val) else ("B0BEC5","37474F"))
            c.fill=PatternFill("solid",fgColor=bg); c.font=Font(bold=True,size=9,color=fg)
            c.alignment=Alignment(horizontal="center",vertical="center")
        elif col==COL_AREA:  # Relevância
            c.fill=PatternFill("solid",fgColor=AREAS_CORES.get(relev,"#94A3B8").lstrip("#"))
            c.font=Font(bold=True,size=9,color="FFFFFF"); c.alignment=Alignment(horizontal="center",vertical="center")
        elif col==COL_TIPO:  # Frente
            c.fill=PatternFill("solid",fgColor=TIPO_CORES_XL.get(frente,"757575"))
            c.font=Font(bold=True,size=9,color="FFFFFF"); c.alignment=Alignment(horizontal="center",vertical="center")
        elif col==COL_ORIGEM:
            c.fill=PatternFill("solid",fgColor="004D40"); c.font=Font(bold=True,size=9,color="FFFFFF")
            c.alignment=Alignment(horizontal="center",vertical="center")
        elif col==COL_PRAZO and val:
            c.fill=PatternFill("solid",fgColor="FFCDD2" if alta else ("FFF9C4" if media else "F5F5F5"))
            c.font=Font(bold=True,size=9,color="B71C1C" if alta else ("E65100" if media else "546E7A"))
            c.alignment=Alignment(horizontal="center",vertical="center")
        elif col==COL_ALTERADOS and val:
            c.fill=PatternFill("solid",fgColor="FFF3E0")
            c.font=Font(italic=True,size=8,color="E65100")
            c.alignment=Alignment(vertical="center",wrap_text=True)
        elif col==COL_TAGS and val:
            # Tags em pílulas — texto em verde quando há projeto prioritário
            tags_str = str(val)
            prioritarias = ("CENTENA","Caldas","Caetité","Santa Quitéria","Angra 3","SMR")
            tem_prio = any(t in tags_str for t in prioritarias)
            c.fill=PatternFill("solid",fgColor="DCFCE7" if tem_prio else "EFF6FF")
            c.font=Font(bold=tem_prio,size=8,color="14532D" if tem_prio else "1E40AF")
            c.alignment=Alignment(vertical="center",wrap_text=True)
        elif col==COL_REVISAO:
            revisao_cores = {
                "✔ Seguido"     : ("1B5E20","FFFFFF"),
                "✘ Não seguido" : ("B71C1C","FFFFFF"),
                "👁 Em análise" : ("1565C0","FFFFFF"),
                "⏸ Aguardando"  : ("546E7A","FFFFFF"),
            }
            bg,fg = revisao_cores.get(str(val),("263548","94A3B8"))
            c.fill=PatternFill("solid",fgColor=bg)
            c.font=Font(bold=True,size=9,color=fg)
            c.alignment=Alignment(horizontal="center",vertical="center")
        elif col==COL_OBSERVACAO and val:
            c.fill=PatternFill("solid",fgColor="E3F2FD")
            c.font=Font(italic=True,size=8,color="1565C0")
            c.alignment=Alignment(vertical="center",wrap_text=True)
        elif alta:
            # Linha 🟢 Alta — destaque verde claro
            c.fill=PatternFill("solid",fgColor="DCFCE7"); c.font=Font(size=9,color="14532D")
        elif media:
            # Linha 🟡 Média — destaque amarelo claro
            c.fill=PatternFill("solid",fgColor="FEF9C3"); c.font=Font(size=9,color="713F12")
        else: c.fill=PatternFill("solid",fgColor=cor_bg)

def aplicar_revisoes_csv(arquivo_excel: str, arquivo_csv: str, registros: dict) -> dict:
    """
    Lê revisoes.csv e sobrescreve as colunas Revisão/Observação/Erro Class./Área Correta
    no Excel e na estrutura registros (usada pelo dashboard).

    Formato do CSV:
        numero,revisao,observacao,erro_classificacao,area_correta

    O CSV é a fonte de verdade para essas 4 colunas — sempre prevalece sobre o Excel.
    Bids no CSV que não existem no Excel são ignorados (sem erro).
    """
    import csv as _csv
    if not Path(arquivo_csv).exists():
        logging.info("revisoes.csv não encontrado — nenhuma revisão a aplicar.")
        return registros

    # Lê CSV
    revisoes = {}
    with open(arquivo_csv, "r", encoding="utf-8", newline="") as f:
        reader = _csv.DictReader(f)
        for r in reader:
            num = (r.get("numero") or "").strip()
            if num:
                revisoes[num] = {
                    "revisao":            (r.get("revisao") or "").strip(),
                    "observacao":         (r.get("observacao") or "").strip(),
                    "erro_classificacao": (r.get("erro_classificacao") or "").strip(),
                    "area_correta":       (r.get("area_correta") or "").strip(),
                }
    print(f"Aplicando {len(revisoes)} revisões do CSV...")
    if not revisoes:
        return registros

    # Atualiza registros em memória
    aplicadas = 0
    for num, rev in revisoes.items():
        if num in registros:
            linha = list(registros[num])
            # Garante tamanho mínimo (cols 18-22)
            while len(linha) < len(COLUNAS):
                linha.append("")
            linha[COL_REVISAO-1]      = rev["revisao"]
            linha[COL_OBSERVACAO-1]   = rev["observacao"]
            linha[COL_ERRO_CLAS-1]    = rev["erro_classificacao"]
            linha[COL_AREA_CORRETA-1] = rev["area_correta"]
            registros[num] = linha
            aplicadas += 1

    # Sobrescreve as células no arquivo Excel
    if not Path(arquivo_excel).exists():
        logging.warning("Excel não existe — pulando atualização de células.")
        return registros

    wb = load_workbook(arquivo_excel)
    if "Base Geral" not in wb.sheetnames:
        return registros
    ws = wb["Base Geral"]

    # Mapeia número -> linha no Excel
    linha_por_numero = {}
    for row_idx in range(4, ws.max_row + 1):
        cel = ws.cell(row=row_idx, column=COL_NUM).value
        if cel:
            linha_por_numero[str(cel).strip()] = row_idx

    for num, rev in revisoes.items():
        if num not in linha_por_numero:
            continue
        row_idx = linha_por_numero[num]
        ws.cell(row=row_idx, column=COL_REVISAO).value      = rev["revisao"]
        ws.cell(row=row_idx, column=COL_OBSERVACAO).value   = rev["observacao"]
        ws.cell(row=row_idx, column=COL_ERRO_CLAS).value    = rev["erro_classificacao"]
        ws.cell(row=row_idx, column=COL_AREA_CORRETA).value = rev["area_correta"]

        # Estilo das novas colunas
        if rev["erro_classificacao"]:
            c = ws.cell(row=row_idx, column=COL_ERRO_CLAS)
            if "Sim" in rev["erro_classificacao"]:
                c.fill = PatternFill("solid", fgColor="B71C1C")
                c.font = Font(bold=True, size=9, color="FFFFFF")
            elif "Não" in rev["erro_classificacao"] or "Nao" in rev["erro_classificacao"]:
                c.fill = PatternFill("solid", fgColor="1B5E20")
                c.font = Font(bold=True, size=9, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb.save(arquivo_excel)
    print(f"  ✓ {aplicadas} revisões aplicadas ao Excel.")
    return registros


def salvar_excel(arquivo,procs_novos=None):
    registros={}; ordem=[]
    if Path(arquivo).exists():
        wb=load_workbook(arquivo)
        if "Base Geral" in wb.sheetnames:
            ws_old=wb["Base Geral"]
            for row in ws_old.iter_rows(min_row=4,values_only=True):
                if row and len(row)>=COL_NUM and row[COL_NUM-1]:
                    num=str(row[COL_NUM-1]).strip()
                    if num not in registros: ordem.append(num)
                    registros[num]=list(row)
            del wb["Base Geral"]
    else:
        wb=Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]

    cont={"novo":0,"atualizado":0,"sem_mudanca":0}

    if procs_novos:
        for p in procs_novos:
            num=str(p.get("numero","")).strip()
            if not num: continue
            nova=montar_linha(p,"")

            if num in registros:
                antiga=registros[num]
                n_ant = len(antiga)
                n_nov = len(nova)


                # Preserva revisão e observação manuais da linha antiga
                while len(nova) < len(COLUNAS): nova.append("")
                if n_ant >= COL_REVISAO:
                    nova[COL_REVISAO-1]   = antiga[COL_REVISAO-1]   or ""
                if n_ant >= COL_OBSERVACAO:
                    nova[COL_OBSERVACAO-1]= antiga[COL_OBSERVACAO-1] or ""

                alterados=[]
                for i in range(1, len(COLUNAS)+1):
                    if i in COLUNAS_IGNORAR_COMPARACAO: continue
                    v_nov = normalizar_val(nova[i-1]   if n_nov >= i else "")
                    v_ant = normalizar_val(antiga[i-1] if n_ant >= i else "")
                    if v_nov != v_ant and (v_nov or v_ant):
                        alterados.append(NOMES_COLUNAS.get(i, f"Col{i}"))

                if alterados:
                    nova[COL_STATUS-1]   = "🔄 Atualizado"
                    nova[COL_ALTERADOS-1]= ", ".join(alterados)
                    cont["atualizado"]  += 1
                else:
                    nova[COL_STATUS-1]   = "✅ Sem mudança"
                    nova[COL_ALTERADOS-1]= ""
                    cont["sem_mudanca"] += 1
            else:
                nova[COL_STATUS-1]   = "🆕 Novo"
                nova[COL_ALTERADOS-1]= ""
                cont["novo"] += 1
                ordem.append(num)
            registros[num]=nova

    ws=wb.create_sheet("Base Geral"); n=len(COLUNAS)
    ws.merge_cells(f"A1:{get_column_letter(n)}1")
    ws["A1"]="CFE MONITOR — Base Geral de Licitações"; hdr(ws["A1"])
    ws.row_dimensions[1].height=30
    ws.merge_cells(f"A2:{get_column_letter(n)}2")
    ws["A2"]=f"Total: {len(registros)} | Atualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font=Font(italic=True,size=9,color="FFFFFF")
    ws["A2"].fill=PatternFill("solid",fgColor=COR_SUBHEAD)
    ws["A2"].alignment=Alignment(horizontal="center"); ws.row_dimensions[2].height=14
    for col,(titulo,larg) in enumerate(COLUNAS,1):
        c=ws.cell(row=3,column=col,value=titulo); chdr(c)
        ws.column_dimensions[get_column_letter(col)].width=larg
    ws.row_dimensions[3].height=22
    prio={"🆕 Novo":0,"🔄 Atualizado":1,"✅ Sem mudança":2}
    nums_ord=sorted(ordem,key=lambda x:(prio.get(str(registros[x][0] if registros.get(x) else ""),3),x))
    row_n=4
    for num in nums_ord:
        ln=registros[num]
        while len(ln)<n: ln.append("")
        estilo_linha(ws,row_n,ln[:n],COR_ALT if (row_n-4)%2==0 else "FFFFFF")
        ws.row_dimensions[row_n].height=30; row_n+=1
    ws.freeze_panes="A4"
    if row_n>4: ws.auto_filter.ref=f"A3:{get_column_letter(n)}{row_n-1}"

    # Dropdown de validação na coluna Revisão
    from openpyxl.worksheet.datavalidation import DataValidation
    opcoes_str = '"' + ','.join(['✔ Seguido','✘ Não seguido','👁 Em análise','⏸ Aguardando']) + '"'
    dv = DataValidation(
        type="list",
        formula1=opcoes_str,
        allow_blank=True,
        showDropDown=False,  # False = mostra o ícone de dropdown na célula
        showErrorMessage=True,
        errorTitle="Opção inválida",
        error="Escolha uma das opções da lista.",
    )
    ws.add_data_validation(dv)
    col_rev = get_column_letter(COL_REVISAO)
    dv.sqref = f"{col_rev}4:{col_rev}{max(row_n, 1000)}" 
    try:
        wb.save(arquivo)
    except PermissionError:
        print(f"\n⚠ ATENÇÃO: Não foi possível salvar '{arquivo}'.")
        print("   Feche o arquivo no Excel e pressione Enter para tentar novamente...")
        input()
        wb.save(arquivo)
        print("   Excel salvo com sucesso!")
    return registros,cont

# ─────────────────────────────────────────────────────────────
#  6. DASHBOARD HTML
# ─────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────
#  6. DASHBOARD HTML
# ─────────────────────────────────────────────────────────────
def gerar_html(registros:dict, arquivo:str):
    print(f"Gerando dashboard HTML: {arquivo}")
    # Schema atual (22 cols): area=Relevância, tipo=Frente, tags=col 22.
    # Mantemos os nomes de chave area/tipo no JS por compatibilidade.
    campos=["status","data","origem","numero","descricao","area","tipo",
            "tipo_proc","contrato","estado","prazo","julgamento","valor",
            "entidade","justificativa","atualizado","campos_alterados",
            "revisao","observacao","erro_clas","relev_correta","tags"]
    dados=[]
    for vals in registros.values():
        while len(vals)<len(campos): vals.append("")
        obj={campos[i]:str(vals[i] or "") for i in range(len(campos))}
        dados.append(obj)

    dados_json         = json.dumps(dados, ensure_ascii=False)
    areas_cores_json   = json.dumps(AREAS_CORES, ensure_ascii=False)
    areas_lista_json   = json.dumps(AREAS_LISTA, ensure_ascii=False)
    # Frentes em formato HTML hex (#xxxxxx) para o JS — TIPO_CORES_XL é sem '#'
    frentes_cores_dict = {f: f"#{c}" for f, c in TIPO_CORES_XL.items()}
    frentes_cores_json = json.dumps(frentes_cores_dict, ensure_ascii=False)
    ts = datetime.now().strftime('%d/%m/%Y %H:%M')

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Rosatom AL — Monitor de Bids</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',sans-serif;background:#0f172a;color:#e2e8f0;min-height:100vh}}
header{{background:linear-gradient(135deg,#1a237e,#283593);padding:18px 32px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 2px 12px #0008}}
header h1{{font-size:1.35rem;font-weight:700}}
.container{{max-width:1700px;margin:0 auto;padding:22px}}

.kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:20px}}
.kpi{{background:#1e293b;border-radius:10px;padding:18px 22px;border-left:4px solid var(--cor)}}
.kpi-label{{font-size:.72rem;text-transform:uppercase;letter-spacing:1px;opacity:.6;margin-bottom:5px}}
.kpi-value{{font-size:2rem;font-weight:800;color:var(--cor)}}
.kpi-sub{{font-size:.72rem;opacity:.45;margin-top:3px}}

/* ── Filtros ── */
.filtros{{background:#1e293b;border-radius:10px;padding:14px 18px;margin-bottom:20px}}
.filtros-row{{display:flex;flex-wrap:wrap;gap:12px;align-items:flex-end}}
.f-group{{display:flex;flex-direction:column;gap:4px}}
.f-group label{{font-size:.7rem;text-transform:uppercase;letter-spacing:.5px;opacity:.55}}
.f-group input[type=date],
.f-group input[type=text]{{background:#0f172a;border:1px solid #334155;color:#e2e8f0;padding:7px 10px;border-radius:7px;font-size:.82rem;width:140px}}
/* Multi-select estilo custom */
.ms-wrap{{position:relative}}
.ms-btn{{background:#0f172a;border:1px solid #334155;color:#e2e8f0;padding:7px 30px 7px 10px;border-radius:7px;font-size:.82rem;min-width:150px;cursor:pointer;display:flex;align-items:center;justify-content:space-between;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.ms-btn:after{{content:'▾';margin-left:6px;opacity:.6;flex-shrink:0}}
.ms-dropdown{{display:none;position:absolute;top:calc(100% + 4px);left:0;background:#1e293b;border:1px solid #334155;border-radius:8px;min-width:190px;max-height:260px;overflow-y:auto;z-index:999;box-shadow:0 8px 24px #0008}}
.ms-dropdown.open{{display:block}}
.ms-item{{display:flex;align-items:center;gap:8px;padding:7px 12px;cursor:pointer;font-size:.82rem}}
.ms-item:hover{{background:#263548}}
.ms-item input[type=checkbox]{{accent-color:#3b82f6;width:14px;height:14px;cursor:pointer}}
.ms-item .dot{{width:10px;height:10px;border-radius:2px;flex-shrink:0}}
.ms-clear{{padding:6px 12px;font-size:.72rem;opacity:.5;border-top:1px solid #334155;cursor:pointer;text-align:right}}
.ms-clear:hover{{opacity:1}}
.btn-reset{{background:#334155;border:none;color:#e2e8f0;padding:8px 14px;border-radius:7px;cursor:pointer;font-size:.82rem}}
.btn-reset:hover{{background:#475569}}

.nuclear-panel{{background:#1a0505;border:1px solid #7f1d1d;border-radius:10px;padding:18px;margin-bottom:20px}}
.nuclear-panel h2{{font-size:.9rem;color:#fca5a5;text-transform:uppercase;letter-spacing:1px;margin-bottom:14px;display:flex;align-items:center;gap:8px}}
.nuclear-breakdown{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}}
.nuc-card{{background:#2d0808;border-radius:8px;padding:14px 16px;border-left:3px solid var(--cor);cursor:pointer;transition:.15s}}
.nuc-card:hover{{background:#3d0a0a;transform:translateY(-1px)}}
.nuc-card.active{{background:#3d0a0a;box-shadow:0 0 0 2px var(--cor)}}
.nuc-card-label{{font-size:.7rem;text-transform:uppercase;letter-spacing:.5px;opacity:.6;margin-bottom:6px;color:#fca5a5}}
.nuc-card-value{{font-size:1.8rem;font-weight:800;color:var(--cor)}}
.nuc-card-sub{{font-size:.7rem;opacity:.45;margin-top:3px;color:#fca5a5}}
.charts{{display:grid;grid-template-columns:2fr 1fr;gap:18px;margin-bottom:20px}}
.chart-box{{background:#1e293b;border-radius:10px;padding:18px}}
.chart-box h3{{font-size:.82rem;opacity:.6;margin-bottom:14px;text-transform:uppercase;letter-spacing:.5px}}
.chart-wrap{{position:relative;height:290px}}

.table-wrap{{background:#1e293b;border-radius:10px;padding:18px;overflow-x:auto}}
.table-wrap h3{{font-size:.82rem;opacity:.6;margin-bottom:10px;text-transform:uppercase;letter-spacing:.5px}}
.table-info{{font-size:.78rem;opacity:.45;margin-bottom:10px}}
table{{width:100%;border-collapse:collapse;font-size:.8rem}}
thead tr{{background:#0f172a}}
th{{padding:9px 11px;text-align:left;font-weight:600;opacity:.65;white-space:nowrap;border-bottom:1px solid #334155;cursor:pointer;user-select:none}}
th:hover{{opacity:1}}
td{{padding:8px 11px;border-bottom:1px solid #1e293b;vertical-align:top}}
tr:hover td{{background:#263548}}
.badge{{display:inline-block;padding:2px 7px;border-radius:20px;font-size:.7rem;font-weight:600;white-space:nowrap}}
.badge-novo{{background:#1b5e20;color:#fff}}
.badge-atualizado{{background:#e65100;color:#fff}}
.badge-sem{{background:#37474f;color:#ccc}}
.area-badge{{display:inline-block;padding:2px 9px;border-radius:4px;font-size:.73rem;font-weight:700;color:#fff;white-space:nowrap}}
.nuclear-row td{{background:#1a0505!important;color:#ffcdd2}}
.nuclear-row:hover td{{background:#2d0808!important}}
.prazo-urgente{{color:#f87171;font-weight:700}}
.prazo-normal{{color:#86efac}}
.paginacao{{display:flex;justify-content:space-between;align-items:center;margin-top:14px;font-size:.8rem;opacity:.55}}
.paginacao button{{background:#334155;border:none;color:#e2e8f0;padding:5px 12px;border-radius:6px;cursor:pointer;margin:0 2px}}
.paginacao button:hover{{background:#475569}}
.paginacao button.active{{background:#3b82f6;color:#fff;opacity:1}}
@media(max-width:900px){{.kpis{{grid-template-columns:repeat(2,1fr)}}.charts{{grid-template-columns:1fr}}}}
</style>
</head>
<body>
<header>
  <div>
    <h1>⚛ Rosatom AL — Monitor de Bids</h1>
    <span id="header-sub">Carregando...</span>
  </div>
  <div style="text-align:right;line-height:1.6">
    <div style="font-size:.7rem;opacity:.45;text-transform:uppercase;letter-spacing:.5px">Última atualização</div>
    <div style="font-size:.92rem;font-weight:600;color:#93c5fd">{ts}</div>
  </div>
</header>
<div class="container">

  <div class="kpis">
    <div class="kpi" style="--cor:#3b82f6"><div class="kpi-label">Total de Licitações</div><div class="kpi-value" id="kpi-total">—</div><div class="kpi-sub" id="kpi-sub-total"></div></div>
    <div class="kpi" style="--cor:#22c55e"><div class="kpi-label">Dias Monitorados</div><div class="kpi-value" id="kpi-dias">—</div><div class="kpi-sub" id="kpi-sub-dias"></div></div>
    <div class="kpi" style="--cor:#a78bfa"><div class="kpi-label">Média / Dia</div><div class="kpi-value" id="kpi-media">—</div><div class="kpi-sub">licitações/dia</div></div>
    <div class="kpi" style="--cor:#22c55e"><div class="kpi-label">🟢 Alta Relevância</div><div class="kpi-value" id="kpi-nuclear">—</div><div class="kpi-sub">disputar — prioridade Rosatom</div></div>
  </div>

  <!-- Painel 🟢 Alta Relevância -->
  <div class="nuclear-panel" id="nuclear-panel" style="background:#022c22;border-color:#14532d">
    <h2 style="color:#86efac">🟢 Painel Alta Relevância — Acompanhamento por Revisão</h2>
    <div class="nuclear-breakdown" id="nuclear-breakdown">
      <div class="nuc-card" style="--cor:#ef4444" onclick="filtrarNuclear('')" id="nuc-pendente">
        <div class="nuc-card-label">🔴 Não revisados</div>
        <div class="nuc-card-value" id="nuc-val-pendente">—</div>
        <div class="nuc-card-sub">Aguardam análise</div>
      </div>
      <div class="nuc-card" style="--cor:#3b82f6" onclick="filtrarNuclear('👁 Em análise')" id="nuc-analise">
        <div class="nuc-card-label">👁 Em análise</div>
        <div class="nuc-card-value" id="nuc-val-analise">—</div>
        <div class="nuc-card-sub">Em avaliação</div>
      </div>
      <div class="nuc-card" style="--cor:#f59e0b" onclick="filtrarNuclear('⏸ Aguardando')" id="nuc-aguardando">
        <div class="nuc-card-label">⏸ Aguardando</div>
        <div class="nuc-card-value" id="nuc-val-aguardando">—</div>
        <div class="nuc-card-sub">Decisão pendente</div>
      </div>
      <div class="nuc-card" style="--cor:#22c55e" onclick="filtrarNuclear('✔ Seguido')" id="nuc-seguido">
        <div class="nuc-card-label">✔ Seguido</div>
        <div class="nuc-card-value" id="nuc-val-seguido">—</div>
        <div class="nuc-card-sub">Concluídos</div>
      </div>
    </div>
    <div style="margin-top:10px;font-size:.72rem;opacity:.4;color:#fca5a5">
      Clique em um card para filtrar a tabela. Os dados abaixo refletem a base completa, independente dos filtros ativos.
    </div>
  </div>

  <div class="filtros">
    <div class="filtros-row">

      <div class="f-group">
        <label>Origem (múltipla)</label>
        <div class="ms-wrap" id="ms-origem">
          <div class="ms-btn" onclick="toggleDropdown('ms-origem')"><span id="ms-origem-label">Todas</span></div>
          <div class="ms-dropdown" id="ms-origem-dd"></div>
        </div>
      </div>

      <div class="f-group">
        <label>Relevância (múltipla)</label>
        <div class="ms-wrap" id="ms-area">
          <div class="ms-btn" onclick="toggleDropdown('ms-area')"><span id="ms-area-label">Todas</span></div>
          <div class="ms-dropdown" id="ms-area-dd"></div>
        </div>
      </div>

      <div class="f-group">
        <label>Frente Rosatom (múltipla)</label>
        <div class="ms-wrap" id="ms-tipo">
          <div class="ms-btn" onclick="toggleDropdown('ms-tipo')"><span id="ms-tipo-label">Todas</span></div>
          <div class="ms-dropdown" id="ms-tipo-dd"></div>
        </div>
      </div>

      <div class="f-group">
        <label>Tag estratégica (múltipla)</label>
        <div class="ms-wrap" id="ms-tags">
          <div class="ms-btn" onclick="toggleDropdown('ms-tags')"><span id="ms-tags-label">Todas</span></div>
          <div class="ms-dropdown" id="ms-tags-dd"></div>
        </div>
      </div>

      <div class="f-group">
        <label>Status (múltiplo)</label>
        <div class="ms-wrap" id="ms-status">
          <div class="ms-btn" onclick="toggleDropdown('ms-status')"><span id="ms-status-label">Todos</span></div>
          <div class="ms-dropdown" id="ms-status-dd"></div>
        </div>
      </div>

      <div class="f-group">
        <label>Mês/Ano (múltiplo)</label>
        <div class="ms-wrap" id="ms-mes">
          <div class="ms-btn" onclick="toggleDropdown('ms-mes')"><span id="ms-mes-label">Todos</span></div>
          <div class="ms-dropdown" id="ms-mes-dd"></div>
        </div>
      </div>

      <div class="f-group">
        <label>Data início</label>
        <input type="date" id="f-data-ini">
      </div>
      <div class="f-group">
        <label>Data fim</label>
        <input type="date" id="f-data-fim">
      </div>

      <div class="f-group">
        <label>Revisão (múltipla)</label>
        <div class="ms-wrap" id="ms-revisao">
          <div class="ms-btn" onclick="toggleDropdown('ms-revisao')"><span id="ms-revisao-label">Todas</span></div>
          <div class="ms-dropdown" id="ms-revisao-dd"></div>
        </div>
      </div>

      <div class="f-group">
        <label>Buscar</label>
        <input type="text" id="f-busca" placeholder="Número ou descrição...">
      </div>
      <div class="f-group">
        <label>&nbsp;</label>
        <button class="btn-reset" onclick="resetarFiltros()">↺ Limpar</button>
      </div>
    </div>
  </div>

  <div class="charts">
    <div class="chart-box"><h3>Licitações por Dia (empilhado por Relevância)</h3><div class="chart-wrap"><canvas id="chart-dias"></canvas></div></div>
    <div class="chart-box"><h3>Distribuição por Frente Rosatom</h3><div class="chart-wrap"><canvas id="chart-areas"></canvas></div></div>
  </div>

  <div class="table-wrap">
    <h3>Registros</h3>
    <div class="table-info" id="table-info"></div>
    <table>
      <thead><tr>
        <th onclick="st('status')">Status ↕</th>
        <th onclick="st('origem')">Origem ↕</th>
        <th onclick="st('data')">Data ↕</th>
        <th onclick="st('numero')">Número ↕</th>
        <th onclick="st('descricao')">Descrição ↕</th>
        <th onclick="st('area')">Relevância ↕</th>
        <th onclick="st('tipo')">Frente ↕</th>
        <th>Tags</th>
        <th onclick="st('prazo')">Prazo Submissão ↕</th>
        <th onclick="st('julgamento')">Julgamento ↕</th>
        <th onclick="st('revisao')">Revisão ↕</th>
        <th onclick="st('campos_alterados')">Campos Alterados ↕</th>
        <th>Observação</th>
        <th>Justificativa</th>
      </tr></thead>
      <tbody id="tbody"></tbody>
    </table>
    <div class="paginacao"><span id="pag-info"></span><div id="pag-btns"></div></div>
  </div>
</div>

<script>
const DADOS_RAW      = {dados_json};
const AREAS_CORES    = {areas_cores_json};   // Relevância → cor (🟢/🟡/🔴)
const AREAS_LISTA    = {areas_lista_json};
const FRENTES_CORES  = {frentes_cores_json}; // Frente → cor (TVEL/ASE/...)
const POR_PAG        = 25;

let dadosFiltrados = [];
let paginaAtual    = 1;
let sortCol        = 'data';
let sortAsc        = false;

// Seleções multi-select
const sel = {{ area: new Set(), tipo: new Set(), tags: new Set(), status: new Set(), origem: new Set(), revisao: new Set(), mes: new Set() }};

// ── Init ────────────────────────────────────────────────────
window.onload = () => {{
  construirMultiSelects();
  document.getElementById('f-data-ini').addEventListener('change', () => {{ paginaAtual=1; aplicarFiltros(); }});
  document.getElementById('f-data-fim').addEventListener('change', () => {{ paginaAtual=1; aplicarFiltros(); }});
  document.getElementById('f-busca').addEventListener('input',  () => {{ paginaAtual=1; aplicarFiltros(); }});
  document.addEventListener('click', e => {{
    if (!e.target.closest('.ms-wrap')) fecharTodos();
  }});
  atualizarPainelNuclear();
  aplicarFiltros();
}};

// ── Multi-select ─────────────────────────────────────────────
function extrairTagsUnicas() {{
  const set = new Set();
  DADOS_RAW.forEach(d => {{
    if (!d.tags) return;
    d.tags.split(',').forEach(t => {{ const x = t.trim(); if (x) set.add(x); }});
  }});
  return [...set].sort();
}}

function construirMultiSelects() {{
  // Áreas = Relevâncias (na ordem 🟢→🟡→🔴, prioridade visual)
  const areas   = AREAS_LISTA.filter(a => DADOS_RAW.some(d=>d.area===a));
  const tipos   = [...new Set(DADOS_RAW.map(d=>d.tipo).filter(Boolean))].sort();
  const tags    = extrairTagsUnicas();
  const statuses= ['🆕 Novo','🔄 Atualizado','✅ Sem mudança'];
  const meses   = [...new Set(DADOS_RAW.map(d=>d.data.substring(0,7)).filter(Boolean))].sort().reverse();
  const mesesLabel = meses.map(m => {{
    const [y,mo]=m.split('-');
    const nome=['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'][parseInt(mo)-1];
    return `${{nome}}/${{y}}`;
  }});

  const origens = [...new Set(DADOS_RAW.map(d=>d.origem).filter(Boolean))].sort();
  criarMS('ms-origem',  'ms-origem-dd',  'ms-origem-label',  origens,  'origem',  null);
  criarMS('ms-area',    'ms-area-dd',    'ms-area-label',    areas,    'area',    AREAS_CORES);
  criarMS('ms-tipo',    'ms-tipo-dd',    'ms-tipo-label',    tipos,    'tipo',    FRENTES_CORES);
  criarMS('ms-tags',    'ms-tags-dd',    'ms-tags-label',    tags,     'tags',    null);
  criarMS('ms-status',  'ms-status-dd',  'ms-status-label',  statuses, 'status',  null);
  criarMS('ms-revisao', 'ms-revisao-dd', 'ms-revisao-label',
    ['✔ Seguido','✘ Não seguido','👁 Em análise','⏸ Aguardando','⬜ Não revisado'], 'revisao', null);
  criarMSComValores('ms-mes', 'ms-mes-dd', 'ms-mes-label', meses, mesesLabel, 'mes');
}}

function criarMSComValores(wrapId, ddId, labelId, valores, labels, selKey) {{
  const dd = document.getElementById(ddId);
  valores.forEach((val, idx) => {{
    const item = document.createElement('div');
    item.className = 'ms-item';
    item.innerHTML = `<input type="checkbox" value="${{val}}" onchange="toggleSel('${{selKey}}','${{val}}','${{labelId}}')">${{labels[idx]}}`;
    dd.appendChild(item);
  }});
  const clr = document.createElement('div');
  clr.className='ms-clear'; clr.textContent='Limpar seleção';
  clr.onclick = () => limparMS(selKey, labelId, ddId);
  dd.appendChild(clr);
}}

function criarMS(wrapId, ddId, labelId, opcoes, selKey, cores) {{
  const dd = document.getElementById(ddId);
  opcoes.forEach(op => {{
    const item = document.createElement('div');
    item.className = 'ms-item';
    const cor = cores ? (cores[op] || '#757575') : null;
    item.innerHTML = `
      <input type="checkbox" value="${{op}}" onchange="toggleSel('${{selKey}}','${{op}}','${{labelId}}')">
      ${{cor ? `<span class="dot" style="background:${{cor}}"></span>` : ''}}
      <span>${{op}}</span>`;
    dd.appendChild(item);
  }});
  const clr = document.createElement('div');
  clr.className='ms-clear'; clr.textContent='Limpar seleção';
  clr.onclick = () => limparMS(selKey, labelId, ddId);
  dd.appendChild(clr);
}}

function toggleSel(key, val, labelId) {{
  if (sel[key].has(val)) sel[key].delete(val); else sel[key].add(val);
  atualizarLabel(key, labelId);
  paginaAtual=1; aplicarFiltros();
}}

function atualizarLabel(key, labelId) {{
  const s = sel[key];
  const defaults = {{ area:'Todas', tipo:'Todas', tags:'Todas', status:'Todos', mes:'Todos', origem:'Todas', revisao:'Todas' }};
  document.getElementById(labelId).textContent =
    s.size === 0 ? (defaults[key] || 'Todos') :
    s.size === 1 ? [...s][0] :
    `${{s.size}} selecionados`;
}}

function limparMS(key, labelId, ddId) {{
  sel[key].clear();
  document.querySelectorAll(`#${{ddId}} input[type=checkbox]`).forEach(cb => cb.checked=false);
  atualizarLabel(key, labelId);
  paginaAtual=1; aplicarFiltros();
}}

function toggleDropdown(wrapId) {{
  const dd = document.getElementById(wrapId+'-dd');
  const wasOpen = dd.classList.contains('open');
  fecharTodos();
  if (!wasOpen) dd.classList.add('open');
}}

function fecharTodos() {{
  document.querySelectorAll('.ms-dropdown').forEach(d => d.classList.remove('open'));
}}

// ── Filtros ──────────────────────────────────────────────────
function resetarFiltros() {{
  ['area','tipo','tags','status','origem','mes','revisao'].forEach(k => {{
    sel[k].clear();
    const ddId = `ms-${{k}}-dd`;
    document.querySelectorAll(`#${{ddId}} input[type=checkbox]`).forEach(cb=>cb.checked=false);
    atualizarLabel(k, `ms-${{k}}-label`);
  }});
  document.getElementById('f-data-ini').value='';
  document.getElementById('f-data-fim').value='';
  document.getElementById('f-busca').value='';
  paginaAtual=1; aplicarFiltros();
}}

function aplicarFiltros() {{
  const dIni  = document.getElementById('f-data-ini').value;
  const dFim  = document.getElementById('f-data-fim').value;
  const busca = document.getElementById('f-busca').value.toLowerCase();

  dadosFiltrados = DADOS_RAW.filter(d => {{
    if (sel.area.size   && !sel.area.has(d.area))              return false;
    if (sel.tipo.size   && !sel.tipo.has(d.tipo))              return false;
    if (sel.tags.size) {{
      const tagList = (d.tags||'').split(',').map(s=>s.trim()).filter(Boolean);
      if (!tagList.some(t=>sel.tags.has(t))) return false;
    }}
    if (sel.status.size && !sel.status.has(d.status))          return false;
    if (sel.mes.size    && !sel.mes.has(d.data.substring(0,7))) return false;
    if (dIni && d.data < dIni) return false;
    if (dFim && d.data > dFim) return false;
    if (sel.origem.size && !sel.origem.has(d.origem)) return false;
    if (sel.revisao.size) {{
      const rev = d.revisao && d.revisao.trim() ? d.revisao : '⬜ Não revisado';
      if (!sel.revisao.has(rev)) return false;
    }}
    if (busca && !d.numero.toLowerCase().includes(busca) && !d.descricao.toLowerCase().includes(busca)) return false;
    return true;
  }});

  dadosFiltrados.sort((a,b) => {{
    let va=a[sortCol]||'', vb=b[sortCol]||'';
    return sortAsc ? va.localeCompare(vb) : vb.localeCompare(va);
  }});

  atualizarKPIs();
  atualizarGraficos();
  renderTabela();
}}

function st(col) {{ if(sortCol===col) sortAsc=!sortAsc; else {{sortCol=col;sortAsc=true;}} aplicarFiltros(); }}

// ── KPIs ────────────────────────────────────────────────────
// ── Painel 🟢 Alta Relevância ─────────────────────────────────
let filtroNuclearAtivo = null;

function atualizarPainelNuclear() {{
  // Sempre usa DADOS_RAW (base completa, não filtrada)
  const altas = DADOS_RAW.filter(d => d.area === '🟢 Alta');
  const pendentes  = altas.filter(d => !d.revisao || !d.revisao.trim()).length;
  const analise    = altas.filter(d => d.revisao === '👁 Em análise').length;
  const aguardando = altas.filter(d => d.revisao === '⏸ Aguardando').length;
  const seguido    = altas.filter(d => d.revisao === '✔ Seguido').length;
  const naoSeguido = altas.filter(d => d.revisao === '✘ Não seguido').length;

  document.getElementById('nuc-val-pendente').textContent   = pendentes;
  document.getElementById('nuc-val-analise').textContent    = analise;
  document.getElementById('nuc-val-aguardando').textContent = aguardando;
  document.getElementById('nuc-val-seguido').textContent    = seguido + (naoSeguido ? ` / ✘ ${{naoSeguido}}` : '');

  // Esconde painel se não há altas
  document.getElementById('nuclear-panel').style.display = altas.length ? 'block' : 'none';
}}

function filtrarNuclear(revisaoVal) {{
  // Toggle: clica de novo para desfiltrar
  if (filtroNuclearAtivo === revisaoVal) {{
    filtroNuclearAtivo = null;
    document.querySelectorAll('.nuc-card').forEach(c => c.classList.remove('active'));
    sel.area.clear(); sel.revisao.clear();
    document.querySelectorAll('#ms-area-dd input, #ms-revisao-dd input').forEach(cb => cb.checked=false);
    atualizarLabel('area','ms-area-label');
    atualizarLabel('revisao','ms-revisao-label');
  }} else {{
    filtroNuclearAtivo = revisaoVal;
    document.querySelectorAll('.nuc-card').forEach(c => c.classList.remove('active'));
    const ids = {{'':'nuc-pendente','👁 Em análise':'nuc-analise','⏸ Aguardando':'nuc-aguardando','✔ Seguido':'nuc-seguido'}};
    if (ids[revisaoVal]) document.getElementById(ids[revisaoVal]).classList.add('active');
    // Filtra por 🟢 Alta + revisão
    sel.area.clear(); sel.area.add('🟢 Alta');
    sel.revisao.clear();
    if (revisaoVal === '') {{
      sel.revisao.add('⬜ Não revisado');
    }} else {{
      sel.revisao.add(revisaoVal);
    }}
    atualizarLabel('area','ms-area-label');
    atualizarLabel('revisao','ms-revisao-label');
  }}
  paginaAtual = 1;
  aplicarFiltros();
}}

function atualizarKPIs() {{
  const total  = dadosFiltrados.length;
  const datas  = new Set(dadosFiltrados.map(d=>d.data).filter(Boolean));
  const dias   = datas.size;
  const media  = dias>0 ? (total/dias).toFixed(1) : '—';
  const nuc    = dadosFiltrados.filter(d=>d.area==='🟢 Alta').length;
  const novos  = dadosFiltrados.filter(d=>d.status.includes('Novo')).length;
  document.getElementById('kpi-total').textContent    = total.toLocaleString();
  document.getElementById('kpi-sub-total').textContent= `${{novos}} novos nesta atualização`;
  document.getElementById('kpi-dias').textContent     = dias;
  document.getElementById('kpi-sub-dias').textContent = dias>0 ? `${{[...datas].sort()[0]}} → ${{[...datas].sort().slice(-1)[0]}}` : '';
  document.getElementById('kpi-media').textContent    = media;
  document.getElementById('kpi-nuclear').textContent  = nuc;
  document.getElementById('header-sub').textContent   = `${{total}} licitações | ${{dias}} dia(s)`;
}}

// ── Gráficos ────────────────────────────────────────────────
let chartDias=null, chartAreas=null;

function atualizarGraficos() {{
  const datas = [...new Set(dadosFiltrados.map(d=>d.data).filter(Boolean))].sort();
  const countPorDia = {{}};
  dadosFiltrados.forEach(d => {{
    if(!d.data) return;
    if(!countPorDia[d.data]) countPorDia[d.data]={{}};
    countPorDia[d.data][d.area] = (countPorDia[d.data][d.area]||0)+1;
  }});

  const datasets = AREAS_LISTA
    .map(area => ({{
      label: area,
      data:  datas.map(d=>(countPorDia[d]||{{}})[area]||0),
      backgroundColor: AREAS_CORES[area],
      borderColor:     AREAS_CORES[area],
      borderWidth: 0,
    }}))
    .filter(ds => ds.data.some(v=>v>0));

  if(chartDias) chartDias.destroy();
  chartDias = new Chart(document.getElementById('chart-dias'), {{
    type:'bar',
    data:{{ labels: datas.map(d => formatDataCompleta(d)), datasets }},
    options:{{
      responsive:true, maintainAspectRatio:false,
      plugins:{{
        legend:{{ position:'bottom', labels:{{ color:'#94a3b8', boxWidth:12, font:{{size:10}} }} }},
        tooltip:{{ mode:'index', intersect:false }}
      }},
      scales:{{
        x:{{ stacked:true, ticks:{{ color:'#94a3b8', maxRotation:45, font:{{size:10}} }}, grid:{{ color:'#1e293b' }} }},
        y:{{ stacked:true, ticks:{{ color:'#94a3b8', stepSize:1 }},               grid:{{ color:'#263548' }} }},
      }}
    }}
  }});

  // Doughnut agora mostra distribuição por FRENTE (campo d.tipo).
  // Excluímos "—" (frente vazia das 🔴 Baixa) para focar nas oportunidades reais.
  const cntFrente={{}};
  dadosFiltrados.forEach(d => {{
    if (!d.tipo || d.tipo === '—') return;
    cntFrente[d.tipo]=(cntFrente[d.tipo]||0)+1;
  }});
  const frentesComDados = Object.entries(cntFrente).filter(([,v])=>v>0).sort((a,b)=>b[1]-a[1]);

  if(chartAreas) chartAreas.destroy();
  chartAreas = new Chart(document.getElementById('chart-areas'), {{
    type:'doughnut',
    data:{{
      labels: frentesComDados.map(([a])=>a),
      datasets:[{{
        data: frentesComDados.map(([,v])=>v),
        backgroundColor: frentesComDados.map(([a])=>FRENTES_CORES[a]||'#757575'),
        borderWidth:2, borderColor:'#1e293b',
      }}]
    }},
    options:{{
      responsive:true, maintainAspectRatio:false,
      plugins:{{
        legend:{{ position:'right', labels:{{ color:'#94a3b8', boxWidth:12, font:{{size:10}} }} }},
        tooltip:{{ callbacks:{{ label: ctx=>`${{ctx.label}}: ${{ctx.raw}} (${{(ctx.raw/dadosFiltrados.length*100).toFixed(1)}}%)` }} }}
      }}
    }}
  }});
}}

// ── Formatação de datas ──────────────────────────────────────
function formatDataCompleta(d) {{
  if(!d) return '';
  const [y,m,dia] = d.split('-');
  return `${{dia}}/${{m}}/${{y}}`;   // dd/mm/aaaa
}}

function formatDataCurta(d) {{
  if(!d) return '';
  const [y,m,dia] = d.split('-');
  return `${{dia}}/${{m}}/${{y.slice(2)}}`;  // dd/mm/aa
}}

// ── Tabela ───────────────────────────────────────────────────
function renderTabela() {{
  const ini=(paginaAtual-1)*POR_PAG, fim=ini+POR_PAG;
  const pagina=dadosFiltrados.slice(ini,fim);
  const tbody=document.getElementById('tbody');
  tbody.innerHTML='';

  pagina.forEach(d => {{
    const tr=document.createElement('tr');
    if(d.area==='🟢 Alta') tr.className='nuclear-row';
    const areaCor   = AREAS_CORES[d.area]||'#94a3b8';
    const frenteCor = FRENTES_CORES[d.tipo]||'#475569';
    const stBadge   = d.status.includes('Novo')?'badge-novo':(d.status.includes('Atualiz')?'badge-atualizado':'badge-sem');
    const prazoC    = isPrazoUrgente(d.prazo)?'prazo-urgente':'prazo-normal';
    const stLabel   = d.status.replace('🆕 ','').replace('🔄 ','').replace('✅ ','');
    const tagsHtml  = (d.tags||'').split(',').map(t=>t.trim()).filter(Boolean)
        .map(t=>{{
          const prio = ['CENTENA','Caldas','Caetité','Santa Quitéria','Angra 3','SMR'].includes(t);
          const bg = prio ? '#15803d' : '#1e3a8a';
          return `<span class="badge" style="background:${{bg}};color:#fff;font-size:.65rem;margin:1px">${{t}}</span>`;
        }}).join('');
    tr.innerHTML=`
      <td><span class="badge ${{stBadge}}">${{stLabel}}</span></td>
      ${{renderOrigem(d.origem)}}
      <td style="white-space:nowrap">${{formatDataCompleta(d.data)}}</td>
      <td style="font-family:monospace;font-size:.76rem;white-space:nowrap">${{d.numero}}</td>
      <td style="min-width:180px;max-width:280px">${{d.descricao}}</td>
      <td><span class="area-badge" style="background:${{areaCor}}">${{d.area}}</span></td>
      <td>${{d.tipo && d.tipo !== '—' ? `<span class="area-badge" style="background:${{frenteCor}};font-size:.7rem">${{d.tipo}}</span>` : '<span style="opacity:.3">—</span>'}}</td>
      <td style="min-width:120px">${{tagsHtml || '<span style="opacity:.3">—</span>'}}</td>
      <td class="${{prazoC}}" style="white-space:nowrap">${{formatDataCompleta(d.prazo)}}</td>
      <td style="font-size:.76rem;opacity:.8;white-space:nowrap">${{formatDataCompleta(d.julgamento)}}</td>
      ${{renderRevisao(d.revisao)}}
      <td style="font-size:.73rem;color:#fb923c;font-style:italic">${{d.campos_alterados||''}}</td>
      <td style="font-size:.75rem;color:#93c5fd;font-style:italic;min-width:160px">${{d.observacao||''}}</td>
      <td style="font-size:.73rem;opacity:.65;min-width:150px">${{d.justificativa}}</td>
    `;
    tbody.appendChild(tr);
  }});

  const total=dadosFiltrados.length, totalPags=Math.ceil(total/POR_PAG);
  document.getElementById('table-info').textContent=`Mostrando ${{ini+1}}–${{Math.min(fim,total)}} de ${{total}} registros`;
  const btns=document.getElementById('pag-btns');
  btns.innerHTML='';
  document.getElementById('pag-info').textContent=`Página ${{paginaAtual}} de ${{totalPags}}`;
  const addBtn=(label,pg,active=false)=>{{
    const b=document.createElement('button');
    b.textContent=label; if(active) b.className='active';
    b.onclick=()=>{{paginaAtual=pg;renderTabela()}};
    btns.appendChild(b);
  }};
  if(paginaAtual>1) addBtn('‹',paginaAtual-1);
  const s=Math.max(1,paginaAtual-2), e=Math.min(totalPags,paginaAtual+2);
  for(let p=s;p<=e;p++) addBtn(p,p,p===paginaAtual);
  if(paginaAtual<totalPags) addBtn('›',paginaAtual+1);
}}

function renderOrigem(orig) {{
  const cfg = {{
    'CFE'          : ['#1565c0','#fff'],
    'Eletronuclear': ['#1b5e20','#fff'],
    'INB'          : ['#6a1b9a','#fff'],
    'CDTN'         : ['#e65100','#fff'],
    'NASA'         : ['#1a237e','#fff'],
    'Dioxitek'     : ['#880e4f','#fff'],
    'CCHEN'        : ['#004d40','#fff'],
    'IAEA'         : ['#37474f','#fff'],
  }};
  const [bg,fg] = cfg[orig] || ['#334155','#94a3b8'];
  return `<td><span class="badge" style="background:${{bg}};color:${{fg}};font-size:.7rem">${{orig||'—'}}</span></td>`;
}}

function renderRevisao(rev) {{
  const cfg = {{
    '✔ Seguido'     : ['#1b5e20','#fff'],
    '✘ Não seguido' : ['#b71c1c','#fff'],
    '👁 Em análise' : ['#1565c0','#fff'],
    '⏸ Aguardando'  : ['#546e7a','#fff'],
  }};
  if (!rev || !cfg[rev]) return `<td style="font-size:.75rem;opacity:.3;text-align:center">—</td>`;
  const [bg,fg] = cfg[rev];
  return `<td><span class="badge" style="background:${{bg}};color:${{fg}}">${{rev}}</span></td>`;
}}

function isPrazoUrgente(prazo) {{
  if(!prazo) return false;
  const hoje=new Date(); hoje.setHours(0,0,0,0);
  const diff=(new Date(prazo)-hoje)/(86400000);
  return diff>=0 && diff<=7;
}}
</script>
</body>
</html>"""

    Path(arquivo).write_text(html, encoding="utf-8")
    print(f"Dashboard HTML salvo: {arquivo}")


# ─────────────────────────────────────────────────────────────
#  7. GITHUB PAGES UPLOAD
# ─────────────────────────────────────────────────────────────
def publicar_github(arquivo_html: str):
    """
    Faz upload do HTML para o GitHub Pages via API REST.
    O arquivo é publicado como index.html na raiz do repositório.
    URL de acesso: https://USUARIO.github.io/REPOSITORIO/
    """
    import base64, urllib.request, urllib.error

    if not GITHUB_TOKEN:
        logging.warning("GITHUB_TOKEN não configurado — pulando publicação no GitHub Pages.")
        return None

    print("Publicando no GitHub Pages...")

    conteudo = Path(arquivo_html).read_bytes()
    conteudo_b64 = base64.b64encode(conteudo).decode()

    api_url = f"https://api.github.com/repos/{GITHUB_USER}/{GITHUB_REPO}/contents/index.html"
    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Content-Type" : "application/json",
        "Accept"       : "application/vnd.github+json",
        "User-Agent"   : "CFE-Monitor/5.0",
    }

    # Verifica se o arquivo já existe (para pegar o SHA necessário para update)
    sha = None
    try:
        req = urllib.request.Request(api_url, headers=headers)
        with urllib.request.urlopen(req) as resp:
            info = json.loads(resp.read())
            sha  = info.get("sha")
    except urllib.error.HTTPError as e:
        if e.code != 404:
            print(f"  Aviso ao verificar arquivo: {e}")

    # Monta payload
    payload = {
        "message": f"Atualização automática — {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "content": conteudo_b64,
        "branch" : GITHUB_BRANCH,
    }
    if sha:
        payload["sha"] = sha  # obrigatório para atualizar arquivo existente

    data = json.dumps(payload).encode()
    req  = urllib.request.Request(api_url, data=data, headers=headers, method="PUT")

    try:
        with urllib.request.urlopen(req) as resp:
            url = f"https://{GITHUB_USER}.github.io/{GITHUB_REPO}/"
            print(f"  ✅ Publicado com sucesso!")
            print(f"  🌐 URL: {url}")
            return url
    except urllib.error.HTTPError as e:
        erro = e.read().decode()
        print(f"  ❌ Erro ao publicar: {e.code} — {erro[:200]}")
        return None


# ─────────────────────────────────────────────────────────────
#  8. MAIN
# ─────────────────────────────────────────────────────────────
def carregar_base_existente(arquivo: str) -> dict:
    """Lê o Excel existente e retorna dict {numero: linha_de_valores}."""
    existentes = {}
    if not Path(arquivo).exists():
        return existentes
    try:
        wb  = load_workbook(arquivo, read_only=True, data_only=True)
        if "Base Geral" not in wb.sheetnames:
            return existentes
        ws  = wb["Base Geral"]
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row and len(row) >= COL_NUM and row[COL_NUM-1]:
                num = str(row[COL_NUM-1]).strip()
                existentes[num] = list(row)
        wb.close()
    except Exception as e:
        print(f"  Aviso: não foi possível ler base existente — {e}")
    return existentes


def _ini_para_fonte(fonte: str, ini_default: str, fim_default: str) -> str:
    """
    Retorna a data inicial efetiva para uma fonte.
    Se a fonte tem lookback configurado, recua a partir de fim_default.
    Caso contrário, retorna ini_default sem alteração.
    """
    lb = LOOKBACK_DIAS_FONTE.get(fonte)
    if lb is None:
        return ini_default
    fim_d = date.fromisoformat(fim_default)
    ini_lookback = (fim_d - timedelta(days=lb)).isoformat()
    # Olha o mais para tras dos dois (lookback ou padrão)
    return ini_lookback if ini_lookback < ini_default else ini_default


def main():
    parser=argparse.ArgumentParser(description="CFE Monitor v5")
    parser.add_argument("--ini",    default=None)
    parser.add_argument("--fim",    default=None)
    parser.add_argument("--excel",  default=ARQUIVO_EXCEL)
    parser.add_argument("--html",   default=ARQUIVO_HTML)
    parser.add_argument("--origem", default=None,
        help="Filtrar por origem: CFE, ETN, INB, CDTN (pode combinar com vírgula: CFE,INB)")
    args=parser.parse_args()

    ini,fim=(args.ini,args.fim) if (args.ini and args.fim) else calcular_intervalo()

    # Valida variáveis de ambiente obrigatórias
    if not os.getenv("ANTHROPIC_API_KEY"):
        logging.error("ANTHROPIC_API_KEY não está configurado. Defina a variável de ambiente.")
        sys.exit(1)
    if not GITHUB_TOKEN:
        logging.warning("GITHUB_TOKEN não configurado — execução continuará mas sem publicar no GitHub Pages.")

    print("="*60)
    print("  CFE Monitor v5")
    print(f"  Período : {ini} → {fim}")
    print(f"  Excel   : {args.excel}")
    print(f"  HTML    : {args.html}")
    print("="*60)

    try:
        # ── 1. Carrega base existente ANTES de qualquer chamada de API ──
        print("Carregando base existente...")
        base_existente = carregar_base_existente(args.excel)
        print(f"  {len(base_existente)} registros na base atual.")


        # ── 2. Define origens ativas ─────────────────────────────────────
        if args.origem:
            origens_ativas = {o.strip().upper() for o in args.origem.split(",")}
        else:
            origens_ativas = {"CFE", "ETN", "INB", "CDTN", "NASA", "DIOXITEK", "CCHEN", "IAEA"}

        session = criar_sessao()

        # CFE
        procs_cfe = []
        if "CFE" in origens_ativas:
            token = obter_token(session)
            procs_cfe = buscar(session, token, ini, fim)
            print(f"CFE: {len(procs_cfe)} procedimentos encontrados.")

        # Eletronuclear via Comprasnet
        procs_elet = []
        if "ETN" in origens_ativas or "ELETRONUCLEAR" in origens_ativas:
            procs_elet = buscar_eletronuclear(session, ini, fim)

        # INB via sistemas.inb.gov.br
        procs_inb = []
        if "INB" in origens_ativas:
            procs_inb = buscar_inb(session, ini, fim)

        # CDTN via Comprasnet (lookback configurável)
        procs_cdtn = []
        if "CDTN" in origens_ativas:
            ini_cdtn = _ini_para_fonte("CDTN", ini, fim)
            procs_cdtn = buscar_cdtn(session, ini_cdtn, fim)

        # NASA — Nucleoeléctrica Argentina
        procs_nasa = []
        if "NASA" in origens_ativas:
            procs_nasa = buscar_nasa(session, ini, fim)

        # Dioxitek — Argentina
        procs_dioxitek = []
        if "DIOXITEK" in origens_ativas:
            procs_dioxitek = buscar_dioxitek(session, ini, fim)

        # CCHEN — Mercado Público Chile (lookback configurável)
        procs_cchen = []
        if "CCHEN" in origens_ativas:
            ini_cchen = _ini_para_fonte("CCHEN", ini, fim)
            procs_cchen = buscar_cchen(session, ini_cchen, fim)

        # IAEA — UNGM (lookback configurável)
        procs_iaea = []
        if "IAEA" in origens_ativas:
            ini_iaea = _ini_para_fonte("IAEA", ini, fim)
            procs_iaea = buscar_iaea(session, ini_iaea, fim)

        procs = procs_cfe + procs_elet + procs_inb + procs_cdtn + procs_nasa + procs_dioxitek + procs_cchen + procs_iaea

        # ── Status por fonte: registra contagem e contador de falhas consecutivas ──
        contagens_fontes = {
            "CFE":      len(procs_cfe),
            "ETN":      len(procs_elet),
            "INB":      len(procs_inb),
            "CDTN":     len(procs_cdtn),
            "NASA":     len(procs_nasa),
            "Dioxitek": len(procs_dioxitek),
            "CCHEN":    len(procs_cchen),
            "IAEA":     len(procs_iaea),
        }
        status_path = Path("status_fontes.json")
        if status_path.exists():
            try:
                status_anterior = json.loads(status_path.read_text(encoding="utf-8"))
            except Exception:
                status_anterior = {}
        else:
            status_anterior = {}
        falhas_ant = status_anterior.get("falhas_consecutivas", {})
        falhas_novo = {}
        for fonte, n in contagens_fontes.items():
            # Só conta como "falha" se a fonte estava ativa nessa execução
            chave_origem = fonte.upper() if fonte != "Dioxitek" else "DIOXITEK"
            if chave_origem not in origens_ativas:
                # Mantém contador anterior se a fonte não foi consultada
                if fonte in falhas_ant:
                    falhas_novo[fonte] = falhas_ant[fonte]
                continue
            if n == 0:
                falhas_novo[fonte] = falhas_ant.get(fonte, 0) + 1
                logging.warning("Fonte %s retornou ZERO procedimentos (falha consecutiva nº %d)", fonte, falhas_novo[fonte])
            else:
                falhas_novo[fonte] = 0
        status_novo = {
            "ultima_execucao_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "periodo": {"ini": ini, "fim": fim},
            "contagens": contagens_fontes,
            "falhas_consecutivas": falhas_novo,
        }
        status_path.write_text(json.dumps(status_novo, indent=2, ensure_ascii=False), encoding="utf-8")

        if not procs:
            print("Nenhum procedimento encontrado.")
            return
        resumo = " + ".join(filter(None, [
            f"{len(procs_cfe)} CFE"   if procs_cfe  else "",
            f"{len(procs_elet)} ETN"  if procs_elet else "",
            f"{len(procs_inb)} INB"   if procs_inb  else "",
            f"{len(procs_cdtn)} CDTN" if procs_cdtn else "",
            f"{len(procs_nasa)} NASA"         if procs_nasa     else "",
            f"{len(procs_dioxitek)} Dioxitek" if procs_dioxitek else "",
            f"{len(procs_cchen)} CCHEN"       if procs_cchen    else "",
            f"{len(procs_iaea)} IAEA"         if procs_iaea     else "",
        ]))
        print(f"Total: {len(procs)} ({resumo})")

        # ── 3. Separa novos dos já existentes ───────────────────────────
        procs_novos     = []
        procs_existentes = []
        for p in procs:
            num = str(p.get("numero","")).strip()
            if num and num in base_existente:
                # Já existe: reaproveita relevância (area)/frente (tipo)/justificativa/tags da base
                linha_ant = base_existente[num]
                p["area"]          = str(linha_ant[COL_AREA-1]  if len(linha_ant)>=COL_AREA  else "🟡 Média")
                p["tipo"]          = str(linha_ant[COL_TIPO-1]  if len(linha_ant)>=COL_TIPO  else "—")
                p["justificativa"] = str(linha_ant[14]          if len(linha_ant)>=15        else "")
                p["tags"]          = str(linha_ant[COL_TAGS-1]  if len(linha_ant)>=COL_TAGS  else "")
                # Preserva data de publicação original para evitar falsos "Atualizado"
                data_orig = str(linha_ant[COL_DATA-1] if len(linha_ant)>=COL_DATA else "").strip()
                if data_orig:
                    p["fecha_pub"] = normalizar_data(data_orig) or p.get("fecha_pub","")
                procs_existentes.append(p)
            else:
                procs_novos.append(p)


        print(f"  🆕 Novos para analisar : {len(procs_novos)}")
        print(f"  ♻  Já na base (skip IA): {len(procs_existentes)}")

        # ── 4. Analisa com Claude APENAS os procedimentos novos ─────────
        if procs_novos:
            procs_novos = analisar(procs_novos)
        else:
            print("Nenhum novo procedimento — Claude não será chamado.")

        # ── 5. Junta tudo e salva ───────────────────────────────────────
        # Inclui registros da base que NÃO vieram na busca atual
        # para que o salvar_excel processe e atualize o status deles
        nums_buscados = {str(p.get("numero","")).strip() for p in procs_novos + procs_existentes}
        procs_base_extras = []
        for num, linha in base_existente.items():
            if num not in nums_buscados:
                # Reconstrói proc mínimo para preservar status/dados
                # Col: 1=Status 2=Data 3=Origem 4=Num 5=Desc 6=Area 7=Tipo
                #      8=TipoProc 9=Contrat 10=Estado 11=Prazo 12=Julg 13=Valor 14=Entidade
                def _g(idx): return str(linha[idx-1] if len(linha)>=idx else "")
                p_extra = {
                    "numero"       : num,
                    "origem"       : _g(3),
                    "descripcion"  : _g(5),
                    "tipo_proc"    : _g(8),
                    "tipo_contrat" : _g(9),
                    "estado"       : _g(10),
                    "fecha_pub"    : normalizar_data(_g(2)),
                    "entidad"      : _g(14),
                    "prazo_sub"    : normalizar_data(_g(11)),
                    "julgamento"   : normalizar_data(_g(12)),
                    "monto"        : _g(13),
                    "area"         : _g(6),    # Relevância (esquema novo)
                    "tipo"         : _g(7),    # Frente (esquema novo)
                    "justificativa": _g(15),
                    "tags"         : _g(COL_TAGS),
                    "id_interno"   : "",
                }
                procs_base_extras.append(p_extra)

        todos = procs_novos + procs_existentes + procs_base_extras
        registros,cont = salvar_excel(args.excel, todos)
        # Aplica revisões manuais do CSV (Revisão/Observação/Erro Class./Área Correta)
        registros = aplicar_revisoes_csv(args.excel, ARQUIVO_REVISOES, registros)
        gerar_html(registros, args.html)
        url_publica = publicar_github(args.html)

        print("\n"+"="*60)
        print("  RESULTADO")
        print("="*60)
        print(f"  🆕 Novos       : {cont['novo']}")
        print(f"  🔄 Atualizados : {cont['atualizado']}")
        print(f"  ✅ Sem mudança : {cont['sem_mudanca']}")
        from collections import Counter
        # Mostra áreas apenas dos novos/atualizados
        # Usa os registros do Excel que têm status gravado
        novos_nums = {str(p.get("numero","")).strip() for p in todos
                      if str(p.get("numero","")).strip() in
                      {num for num,ln in registros.items()
                       if str(ln[COL_STATUS-1] if len(ln)>=COL_STATUS else "").startswith(("🆕","🔄"))}}
        areas_base = [p for p in todos if str(p.get("numero","")).strip() in novos_nums] or todos
        label = "novos/atualizados" if novos_nums else "total"
        # area = Relevância (🟢 Alta / 🟡 Média / 🔴 Baixa)
        areas = Counter(p.get("area","🟡 Média") for p in areas_base)
        print()
        print(f"  Relevância ({label}):")
        for area,qtd in sorted(areas.items(), key=lambda x:-x[1]):
            print(f"    {area:<22}: {qtd}{' ⚠' if area=='🟢 Alta' else ''}")

        print(f"\n  Excel : {args.excel}")
        print(f"  HTML  : {args.html}")
        if url_publica:
            print(f"  🌐 Online: {url_publica}")
        print("="*60)

    except requests.exceptions.ConnectionError: print("ERRO: Falha de conexão.")
    except Exception as e: print(f"ERRO: {e}"); raise

if __name__=="__main__":
    main()
